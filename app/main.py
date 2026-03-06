import json
import os
import re
import shutil
import subprocess
import sys
import uuid
import wave
from io import BytesIO
from datetime import datetime, timezone
from pathlib import Path

import pyttsx3

from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from google.auth.transport.requests import Request as GoogleRequest
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import Flow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaFileUpload

# Allow local OAuth redirect over HTTP during development.
os.environ["OAUTHLIB_INSECURE_TRANSPORT"] = "1"

BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "data"
UPLOAD_DIR = DATA_DIR / "uploads"
OUTPUT_DIR = DATA_DIR / "outputs"
CREDENTIALS_DIR = DATA_DIR / "credentials"
VIDEO_LIBRARY_DIR = DATA_DIR / "video_library"
VIDEO_LIBRARY_CATALOG = VIDEO_LIBRARY_DIR / "catalog.json"
VIDEO_LIBRARY_INDEX = VIDEO_LIBRARY_DIR / "video_index.json"
AUDIO_LIBRARY_DIR = DATA_DIR / "audio_library"
AUDIO_LIBRARY_CATALOG = AUDIO_LIBRARY_DIR / "catalog.json"
SCRIPTS_DIR = DATA_DIR / "scripts"
MUSIC_DIR = DATA_DIR / "music"
LOGOS_DIR = DATA_DIR / "logos"
THUMBNAILS_DIR = DATA_DIR / "thumbnails"
BRANDING_FILE = DATA_DIR / "branding_packs.json"
VOICES_DIR = DATA_DIR / "voices"
PIPER_VOICES_DIR = VOICES_DIR / "piper"
TOKEN_PATH = CREDENTIALS_DIR / "token.json"
CLIENT_SECRET_PATH = CREDENTIALS_DIR / "client_secret.json"

UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
CREDENTIALS_DIR.mkdir(parents=True, exist_ok=True)
VIDEO_LIBRARY_DIR.mkdir(parents=True, exist_ok=True)
AUDIO_LIBRARY_DIR.mkdir(parents=True, exist_ok=True)
SCRIPTS_DIR.mkdir(parents=True, exist_ok=True)
MUSIC_DIR.mkdir(parents=True, exist_ok=True)
LOGOS_DIR.mkdir(parents=True, exist_ok=True)
THUMBNAILS_DIR.mkdir(parents=True, exist_ok=True)
PIPER_VOICES_DIR.mkdir(parents=True, exist_ok=True)

SCOPES = ["https://www.googleapis.com/auth/youtube.upload"]
VOICE_STYLE_PRESETS = {
    "professional": {"rate_adjust": -5},
    "casual": {"rate_adjust": 5},
    "narrator": {"rate_adjust": -10},
    "energetic": {"rate_adjust": 15},
    "calm": {"rate_adjust": -15},
    "dramatic": {"rate_adjust": -8},
}
SUBTITLE_STYLE_PRESETS = {
    "classic": {
        "text_color": "#FFFFFF",
        "bg_color": "#000000",
        "bold": False,
        "italic": False,
        "alignment": 2,
        "font_size": 48,
        "font_name": "Arial",
        "border_style": 3,
        "outline": 1,
        "shadow": 0,
        "margin_v": 80,
    },
    "viral": {
        "text_color": "#FFFFFF",
        "bg_color": "#FF4D00",
        "bold": True,
        "italic": False,
        "alignment": 2,
        "font_size": 56,
        "font_name": "Arial",
        "border_style": 3,
        "outline": 1,
        "shadow": 0,
        "margin_v": 82,
    },
    "reels": {
        "text_color": "#F5F7FA",
        "bg_color": "#1F2937",
        "bold": True,
        "italic": False,
        "alignment": 2,
        "font_size": 50,
        "font_name": "Arial",
        "border_style": 3,
        "outline": 1,
        "shadow": 0,
        "margin_v": 84,
    },
    "cinematic": {
        "text_color": "#F7E7C6",
        "bg_color": "#101010",
        "bold": False,
        "italic": False,
        "alignment": 2,
        "font_size": 46,
        "font_name": "Arial",
        "border_style": 3,
        "outline": 1,
        "shadow": 0,
        "margin_v": 96,
    },
}
OUTPUT_MODE_PRESETS = {
    "youtube": {"width": 1920, "height": 1080},
    "shorts": {"width": 1080, "height": 1920},
    "reels": {"width": 1080, "height": 1920},
    "square": {"width": 1080, "height": 1080},
}
SUBTITLE_ASS_TEMPLATES = {
    "fade": "fade",
    "bold_center": "bold_center",
    "karaoke_word_by_word": "karaoke_word_by_word",
    "bounce_fade": "bounce_fade",
    "beat_sync": "beat_sync",
}
VOICE_MODEL_MATRIX = {
    "professional": {
        "male": "en_US-john-medium.onnx",
        "female": "en_US-kristin-medium.onnx",
    },
    "casual": {
        "male": "en_US-joe-medium.onnx",
        "female": "en_US-amy-medium.onnx",
    },
    "narrator": {
        "male": "en_US-norman-medium.onnx",
        "female": "en_US-libritts-high.onnx",
    },
    "energetic": {
        "male": "en_US-sam-medium.onnx",
        "female": "en_US-hfc_female-medium.onnx",
    },
    "calm": {
        "male": "en_US-hfc_male-medium.onnx",
        "female": "en_US-kathleen-low.onnx",
    },
    "dramatic": {
        "male": "en_US-ryan-high.onnx",
        "female": "en_US-ljspeech-high.onnx",
    },
}
VOICE_STYLE_SOUND_PROFILES = {
    "professional": {
        "length_scale": 1.00,
        "noise_scale": 0.40,
        "noise_w_scale": 0.70,
        "volume": 1.00,
        "tempo": 1.00,
        "bass_gain": 1.0,
        "presence_gain": 0.5,
    },
    "casual": {
        "length_scale": 0.92,
        "noise_scale": 0.55,
        "noise_w_scale": 0.85,
        "volume": 1.03,
        "tempo": 1.08,
        "bass_gain": 0.5,
        "presence_gain": 1.5,
    },
    "narrator": {
        "length_scale": 1.08,
        "noise_scale": 0.35,
        "noise_w_scale": 0.65,
        "volume": 1.02,
        "tempo": 0.92,
        "bass_gain": 2.5,
        "presence_gain": 0.0,
    },
    "energetic": {
        "length_scale": 0.86,
        "noise_scale": 0.65,
        "noise_w_scale": 0.90,
        "volume": 1.10,
        "tempo": 1.15,
        "bass_gain": 0.5,
        "presence_gain": 3.0,
    },
    "calm": {
        "length_scale": 1.15,
        "noise_scale": 0.30,
        "noise_w_scale": 0.60,
        "volume": 0.95,
        "tempo": 0.88,
        "bass_gain": 0.0,
        "presence_gain": -2.0,
    },
    "dramatic": {
        "length_scale": 1.12,
        "noise_scale": 0.50,
        "noise_w_scale": 0.80,
        "volume": 1.10,
        "tempo": 0.90,
        "bass_gain": 4.0,
        "presence_gain": 2.0,
    },
}
CONTEXT_CATEGORY_RULES = {
    "motivation": {
        "tokens": {
            "motivation",
            "motivational",
            "discipline",
            "consistent",
            "consistency",
            "grind",
            "hustle",
            "effort",
            "focus",
            "persist",
            "persevere",
        },
        "phrases": {"never give up", "keep going", "stay focused"},
    },
    "mindset": {
        "tokens": {
            "mindset",
            "belief",
            "confidence",
            "mentality",
            "psychology",
            "thinking",
            "perspective",
            "clarity",
            "attitude",
            "selftalk",
        },
        "phrases": {"growth mindset", "mental model", "change your thinking"},
    },
    "success": {
        "tokens": {
            "success",
            "achieve",
            "achievement",
            "winner",
            "winning",
            "goal",
            "results",
            "milestone",
            "accomplish",
            "top",
        },
        "phrases": {"path to success", "achieve your goal", "become successful"},
    },
    "business": {
        "tokens": {
            "business",
            "startup",
            "entrepreneur",
            "company",
            "market",
            "strategy",
            "team",
            "office",
            "management",
            "sales",
            "brand",
        },
        "phrases": {"business model", "start a business", "build a company"},
    },
    "finance": {
        "tokens": {
            "finance",
            "money",
            "investment",
            "invest",
            "wealth",
            "income",
            "savings",
            "budget",
            "stock",
            "trading",
            "economy",
            "debt",
            "profit",
        },
        "phrases": {"financial freedom", "passive income", "stock market"},
    },
    "politics": {
        "tokens": {
            "politics",
            "political",
            "election",
            "government",
            "minister",
            "parliament",
            "policy",
            "senate",
            "democracy",
            "campaign",
            "vote",
            "party",
        },
        "phrases": {"public policy", "general election", "political campaign"},
    },
    "currentaffairs": {
        "tokens": {
            "news",
            "breaking",
            "update",
            "current",
            "affairs",
            "headline",
            "international",
            "global",
            "today",
            "latest",
            "report",
            "crisis",
        },
        "phrases": {"breaking news", "latest update", "current affairs"},
    },
    "socialcommentary": {
        "tokens": {
            "society",
            "social",
            "culture",
            "community",
            "issue",
            "debate",
            "opinion",
            "trend",
            "public",
            "narrative",
            "commentary",
        },
        "phrases": {"social issue", "public opinion", "cultural shift"},
    },
    "reallifestories": {
        "tokens": {
            "story",
            "real",
            "life",
            "journey",
            "family",
            "struggle",
            "experience",
            "incident",
            "lesson",
            "human",
            "person",
        },
        "phrases": {"real life story", "true story", "life lesson"},
    },
    "viraltrends": {
        "tokens": {
            "viral",
            "trend",
            "trending",
            "shorts",
            "reel",
            "reels",
            "socialmedia",
            "internet",
            "meme",
            "popular",
            "algorithm",
            "creator",
        },
        "phrases": {"viral trend", "going viral", "social media trend"},
    },
}


def _voice_profile(voice_style: str, rate: int) -> dict[str, float]:
    profile = VOICE_STYLE_SOUND_PROFILES.get(
        voice_style, VOICE_STYLE_SOUND_PROFILES["professional"]
    ).copy()
    rate_factor = 175 / max(120, min(260, rate))
    profile["length_scale"] = max(0.70, min(1.45, profile["length_scale"] * rate_factor))
    profile["tempo"] = max(0.75, min(1.25, profile["tempo"]))
    return profile


def _resolve_subtitle_style_preset(preset_name: str | None) -> dict:
    preset_key = str(preset_name or "classic").strip().lower()
    if preset_key == "default":
        preset_key = "classic"
    preset = SUBTITLE_STYLE_PRESETS.get(preset_key)
    if not preset:
        preset = SUBTITLE_STYLE_PRESETS["classic"]
    return preset.copy()


def _resolve_subtitle_template(template_name: str | None) -> str:
    template_key = str(template_name or "fade").strip().lower()
    aliases = {
        "default": "fade",
        "fade_only": "fade",
        "boldcenter": "bold_center",
        "bold_center_captions": "bold_center",
        "bold center": "bold_center",
        "center bold": "bold_center",
        "karaoke": "karaoke_word_by_word",
        "word_by_word": "karaoke_word_by_word",
        "word by word": "karaoke_word_by_word",
        "karaoke_word": "karaoke_word_by_word",
        "bounce": "bounce_fade",
        "bouncefade": "bounce_fade",
        "bounce/fade": "bounce_fade",
        "bounce fade": "bounce_fade",
        "beat": "beat_sync",
        "beat_sync": "beat_sync",
        "beat sync": "beat_sync",
        "rhythm": "beat_sync",
        "rhythm_sync": "beat_sync",
        "rhythm sync": "beat_sync",
    }
    template_key = aliases.get(template_key, template_key)
    if template_key not in SUBTITLE_ASS_TEMPLATES:
        template_key = "fade"
    return template_key


def _resolve_output_mode(output_mode: str | None) -> str:
    mode = str(output_mode or "youtube").strip().lower()
    aliases = {
        "16:9": "youtube",
        "landscape": "youtube",
        "standard": "youtube",
        "youtube_standard": "youtube",
        "9:16": "shorts",
        "vertical": "shorts",
        "yt_shorts": "shorts",
        "short": "shorts",
        "reel": "reels",
        "1:1": "square",
        "instagram_square": "square",
    }
    mode = aliases.get(mode, mode)
    if mode not in OUTPUT_MODE_PRESETS:
        mode = "youtube"
    return mode


def _output_mode_resolution(output_mode: str | None) -> tuple[int, int]:
    mode = _resolve_output_mode(output_mode)
    preset = OUTPUT_MODE_PRESETS[mode]
    return int(preset["width"]), int(preset["height"])


def _subtitle_margin_for_mode(
    base_margin_v: int, output_mode: str, play_res_y: int
) -> int:
    mode = _resolve_output_mode(output_mode)
    scale_factor = max(0.8, play_res_y / 720.0)
    mode_boost = {
        "youtube": 1.00,
        "square": 1.15,
        "shorts": 1.40,
        "reels": 1.40,
    }.get(mode, 1.0)
    return max(40, int(round(base_margin_v * scale_factor * mode_boost)))


def _post_process_voice(audio_path: Path, profile: dict[str, float]) -> None:
    processed_path = audio_path.with_name(f"{audio_path.stem}_fx.wav")
    filters = [
        f"atempo={profile['tempo']:.2f}",
        f"equalizer=f=120:t=o:w=1.2:g={profile['bass_gain']:.1f}",
        f"equalizer=f=3200:t=o:w=1.1:g={profile['presence_gain']:.1f}",
        "acompressor=threshold=-18dB:ratio=2.8:attack=18:release=150",
        f"volume={profile['volume']:.2f}",
    ]
    command = [
        "ffmpeg",
        "-y",
        "-i",
        str(audio_path),
        "-filter:a",
        ",".join(filters),
        str(processed_path),
    ]
    process = subprocess.run(command, capture_output=True, text=True)
    if process.returncode == 0 and processed_path.exists():
        processed_path.replace(audio_path)
app = FastAPI()
app.mount("/static", StaticFiles(directory=BASE_DIR / "app" / "static"), name="static")

templates = Jinja2Templates(directory=BASE_DIR / "app" / "templates")


@app.exception_handler(RuntimeError)
async def runtime_error_handler(request: Request, exc: RuntimeError) -> JSONResponse:
    return JSONResponse({"error": str(exc)}, status_code=500)


@app.exception_handler(Exception)
async def generic_error_handler(request: Request, exc: Exception) -> JSONResponse:
    return JSONResponse({"error": str(exc)}, status_code=500)


@app.get("/", response_class=HTMLResponse)
def index(request: Request) -> HTMLResponse:
    return templates.TemplateResponse("index.html", {"request": request})


@app.get("/api/library")
def library_catalog() -> JSONResponse:
    return JSONResponse({"videos": _load_video_catalog()})


@app.get("/api/library/categories")
def library_categories() -> JSONResponse:
    return JSONResponse({"categories": _available_context_categories()})


@app.get("/api/audio-library")
def audio_library_catalog() -> JSONResponse:
    return JSONResponse({"audio": _load_audio_catalog()})


@app.get("/api/branding-packs")
def branding_packs() -> JSONResponse:
    return JSONResponse({"packs": _load_branding_packs()})


def _run_ffmpeg(command: list[str]) -> None:
    process = subprocess.run(command, capture_output=True, text=True)
    if process.returncode != 0:
        raise RuntimeError(f"FFmpeg failed: {process.stderr}")


def _safe_filename(name: str) -> str:
    return re.sub(r"[^a-zA-Z0-9._-]", "_", name)


def _generate_thumbnail_from_video(
    video_path: Path, thumbnail_path: Path, seek_seconds: float = 0.6
) -> None:
    command = [
        "ffmpeg",
        "-y",
        "-ss",
        f"{max(0.0, float(seek_seconds)):.2f}",
        "-i",
        str(video_path),
        "-frames:v",
        "1",
        "-q:v",
        "2",
        str(thumbnail_path),
    ]
    _run_ffmpeg(command)


def _apply_logo_overlay_to_video(
    video_path: Path,
    logo_path: Path,
    output_path: Path,
    logo_position: str,
    logo_scale_percent: int,
    logo_animated: bool = False,
) -> None:
    position_map = {
        "top-left": "20:20",
        "top-right": "W-w-20:20",
        "bottom-left": "20:H-h-20",
        "bottom-right": "W-w-20:H-h-20",
        "center": "(W-w)/2:(H-h)/2",
    }
    overlay_expr = position_map.get(logo_position, position_map["top-right"])
    scale_factor = max(5, min(40, logo_scale_percent)) / 100.0
    if logo_animated:
        filter_complex = (
            f"[1:v][0:v]scale2ref=w=iw*{scale_factor:.4f}:h=ow/mdar[logo][base];"
            "[logo]format=rgba,fade=t=in:st=0:d=1.2:alpha=1[logoa];"
            f"[base][logoa]overlay={overlay_expr}[v]"
        )
    else:
        filter_complex = (
            f"[1:v][0:v]scale2ref=w=iw*{scale_factor:.4f}:h=ow/mdar[logo][base];"
            f"[base][logo]overlay={overlay_expr}[v]"
        )
    command = [
        "ffmpeg",
        "-y",
        "-i",
        str(video_path),
        "-i",
        str(logo_path),
        "-filter_complex",
        filter_complex,
        "-map",
        "[v]",
        "-map",
        "0:a?",
        "-c:v",
        "libx264",
        "-c:a",
        "copy",
        "-movflags",
        "+faststart",
        str(output_path),
    ]
    _run_ffmpeg(command)


def _video_duration_seconds(video_path: Path) -> float:
    command = [
        "ffprobe",
        "-v",
        "error",
        "-show_entries",
        "format=duration",
        "-of",
        "default=noprint_wrappers=1:nokey=1",
        str(video_path),
    ]
    process = subprocess.run(command, capture_output=True, text=True)
    if process.returncode != 0:
        raise RuntimeError(f"FFprobe failed: {process.stderr}")
    try:
        return float((process.stdout or "").strip())
    except ValueError as exc:
        raise RuntimeError("Unable to parse video duration.") from exc


def _video_dimensions(video_path: Path) -> tuple[int, int]:
    command = [
        "ffprobe",
        "-v",
        "error",
        "-select_streams",
        "v:0",
        "-show_entries",
        "stream=width,height",
        "-of",
        "csv=p=0:s=x",
        str(video_path),
    ]
    process = subprocess.run(command, capture_output=True, text=True)
    if process.returncode != 0:
        raise RuntimeError(f"FFprobe failed: {process.stderr}")
    dimensions = (process.stdout or "").strip().split("x")
    if len(dimensions) != 2:
        raise RuntimeError("Unable to parse video dimensions.")
    try:
        return int(dimensions[0]), int(dimensions[1])
    except ValueError as exc:
        raise RuntimeError("Unable to parse video dimensions.") from exc


def _build_smart_framing_filter(video_path: Path, output_mode: str) -> str:
    target_w, target_h = _output_mode_resolution(output_mode)
    source_w, source_h = _video_dimensions(video_path)
    if source_w <= 0 or source_h <= 0:
        return f"scale={target_w}:{target_h}:flags=lanczos,setsar=1"

    source_aspect = source_w / source_h
    target_aspect = target_w / target_h

    if abs(source_aspect - target_aspect) < 0.01:
        return f"scale={target_w}:{target_h}:flags=lanczos,setsar=1"

    if source_aspect > target_aspect:
        # Source is wider than target frame: scale to target height and center-crop.
        return (
            f"scale=-2:{target_h}:flags=lanczos,"
            f"crop={target_w}:{target_h}:(iw-{target_w})/2:(ih-{target_h})/2,setsar=1"
        )

    # Source is narrower/taller than target frame: scale to target height and pad.
    return (
        f"scale=-2:{target_h}:flags=lanczos,"
        f"pad={target_w}:{target_h}:(ow-iw)/2:(oh-ih)/2:black,setsar=1"
    )


def _ffmpeg_drawtext_escape(text: str) -> str:
    escaped = text.replace("\\", "\\\\")
    escaped = escaped.replace(":", r"\:")
    escaped = escaped.replace("'", r"\'")
    escaped = escaped.replace("%", r"\%")
    escaped = escaped.replace("\n", r"\n")
    return escaped


def _apply_end_credits_to_video(
    video_path: Path,
    output_path: Path,
    credits_text: str,
    credits_duration_sec: int,
) -> None:
    duration = _video_duration_seconds(video_path)
    credits_window = max(2, min(30, int(credits_duration_sec)))
    start_at = max(0.0, duration - float(credits_window))
    escaped_text = _ffmpeg_drawtext_escape(credits_text.strip())
    filter_chain = (
        f"drawbox=x=0:y=ih-240:w=iw:h=240:color=black@0.70:t=fill:enable='gte(t,{start_at:.3f})',"
        f"drawtext=text='{escaped_text}':fontcolor=white:fontsize=44:"
        f"x=(w-text_w)/2:y=h-170:line_spacing=12:enable='gte(t,{start_at:.3f})'"
    )
    command = [
        "ffmpeg",
        "-y",
        "-i",
        str(video_path),
        "-vf",
        filter_chain,
        "-c:v",
        "libx264",
        "-c:a",
        "copy",
        "-movflags",
        "+faststart",
        str(output_path),
    ]
    _run_ffmpeg(command)


def _apply_branding_overlays_to_video(
    video_path: Path,
    output_path: Path,
    *,
    intro_text: str = "",
    intro_duration_sec: int = 0,
    outro_text: str = "",
    outro_duration_sec: int = 0,
    subscribe_cta_text: str = "",
    subscribe_cta_duration_sec: int = 5,
    subscribe_cta_from_end_sec: int = 12,
    end_screen_blocks: bool = False,
    end_screen_duration_sec: int = 8,
) -> None:
    duration = _video_duration_seconds(video_path)
    vf_parts: list[str] = []

    intro_text_value = str(intro_text or "").strip()
    if intro_text_value and int(intro_duration_sec or 0) > 0:
        intro_seconds = max(1, min(15, int(intro_duration_sec)))
        escaped_intro = _ffmpeg_drawtext_escape(intro_text_value)
        vf_parts.append(
            "drawbox=x=0:y=0:w=iw:h=ih:color=black@0.40:t=fill:"
            f"enable='between(t,0,{float(intro_seconds):.3f})'"
        )
        vf_parts.append(
            f"drawtext=text='{escaped_intro}':fontcolor=white:fontsize=54:"
            f"x=(w-text_w)/2:y=h*0.12:enable='between(t,0,{float(intro_seconds):.3f})'"
        )

    cta_text = str(subscribe_cta_text or "").strip()
    if cta_text:
        cta_duration = max(2, min(20, int(subscribe_cta_duration_sec or 5)))
        cta_from_end = max(2, min(120, int(subscribe_cta_from_end_sec or 12)))
        cta_start = max(0.0, duration - float(cta_from_end))
        cta_end = min(duration, cta_start + float(cta_duration))
        escaped_cta = _ffmpeg_drawtext_escape(cta_text)
        vf_parts.append(
            "drawbox=x=(w*0.22):y=(h*0.78):w=(w*0.56):h=84:color=red@0.75:t=fill:"
            f"enable='between(t,{cta_start:.3f},{cta_end:.3f})'"
        )
        vf_parts.append(
            f"drawtext=text='{escaped_cta}':fontcolor=white:fontsize=42:"
            f"x=(w-text_w)/2:y=h*0.81:enable='between(t,{cta_start:.3f},{cta_end:.3f})'"
        )

    end_duration = max(2, min(20, int(end_screen_duration_sec or 8)))
    end_start = max(0.0, duration - float(end_duration))
    if _to_bool(end_screen_blocks, default=False):
        vf_parts.append(
            "drawbox=x=0:y=0:w=iw:h=ih:color=black@0.20:t=fill:"
            f"enable='gte(t,{end_start:.3f})'"
        )
        vf_parts.append(
            "drawbox=x=w*0.08:y=h*0.18:w=w*0.38:h=h*0.42:color=white@0.20:t=3:"
            f"enable='gte(t,{end_start:.3f})'"
        )
        vf_parts.append(
            "drawbox=x=w*0.54:y=h*0.18:w=w*0.38:h=h*0.42:color=white@0.20:t=3:"
            f"enable='gte(t,{end_start:.3f})'"
        )
        vf_parts.append(
            "drawbox=x=w*0.32:y=h*0.68:w=w*0.36:h=h*0.20:color=white@0.20:t=3:"
            f"enable='gte(t,{end_start:.3f})'"
        )
        vf_parts.append(
            "drawtext=text='Watch Next':fontcolor=white:fontsize=30:"
            f"x=w*0.16:y=h*0.22:enable='gte(t,{end_start:.3f})'"
        )
        vf_parts.append(
            "drawtext=text='Recommended':fontcolor=white:fontsize=30:"
            f"x=w*0.60:y=h*0.22:enable='gte(t,{end_start:.3f})'"
        )
        vf_parts.append(
            "drawtext=text='SUBSCRIBE':fontcolor=white:fontsize=34:"
            f"x=(w-text_w)/2:y=h*0.75:enable='gte(t,{end_start:.3f})'"
        )

    outro_text_value = str(outro_text or "").strip()
    if outro_text_value and int(outro_duration_sec or 0) > 0:
        outro_seconds = max(1, min(20, int(outro_duration_sec)))
        outro_start = max(0.0, duration - float(outro_seconds))
        escaped_outro = _ffmpeg_drawtext_escape(outro_text_value)
        vf_parts.append(
            "drawbox=x=0:y=0:w=iw:h=ih:color=black@0.50:t=fill:"
            f"enable='between(t,{outro_start:.3f},{duration:.3f})'"
        )
        vf_parts.append(
            f"drawtext=text='{escaped_outro}':fontcolor=white:fontsize=52:"
            f"x=(w-text_w)/2:y=h*0.10:enable='between(t,{outro_start:.3f},{duration:.3f})'"
        )

    if not vf_parts:
        shutil.copyfile(video_path, output_path)
        return
    command = [
        "ffmpeg",
        "-y",
        "-i",
        str(video_path),
        "-vf",
        ",".join(vf_parts),
        "-c:v",
        "libx264",
        "-c:a",
        "copy",
        "-movflags",
        "+faststart",
        str(output_path),
    ]
    _run_ffmpeg(command)


def _prepare_video_for_shorts(
    video_path: Path,
    output_path: Path,
    max_duration_sec: int = 59,
) -> None:
    filter_chain = (
        "scale=1080:1920:force_original_aspect_ratio=decrease,"
        "pad=1080:1920:(ow-iw)/2:(oh-ih)/2:black,setsar=1"
    )
    command = [
        "ffmpeg",
        "-y",
        "-i",
        str(video_path),
        "-t",
        str(max_duration_sec),
        "-vf",
        filter_chain,
        "-c:v",
        "libx264",
        "-c:a",
        "aac",
        "-movflags",
        "+faststart",
        str(output_path),
    ]
    _run_ffmpeg(command)


def _ffmpeg_filter_path(path: Path) -> str:
    normalized = path.as_posix()
    return (
        normalized.replace("\\", "/")
        .replace(":", "\\:")
        .replace("'", "\\'")
        .replace(" ", "\\ ")
    )


def _ffmpeg_subtitles_path(path: Path) -> str:
    try:
        relative = path.relative_to(BASE_DIR)
    except ValueError:
        relative = path.resolve()
    return _ffmpeg_filter_path(relative)


def _parse_publish_at(publish_at: str) -> str | None:
    """Parse publish_at to RFC 3339 UTC string for YouTube.

    Returns None when the value is blank, invalid, or in the past.
    """
    if not publish_at or not str(publish_at).strip():
        return None
    raw_value = str(publish_at).strip()
    try:
        # Accept ISO values with or without timezone.
        dt = datetime.fromisoformat(raw_value.replace("Z", "+00:00"))
        if dt.tzinfo is None:
            # For naive datetimes (common from Excel), assume local system timezone.
            local_tz = datetime.now().astimezone().tzinfo or timezone.utc
            dt = dt.replace(tzinfo=local_tz)
        utc_value = dt.astimezone(timezone.utc)
        if utc_value <= datetime.now(timezone.utc):
            return None
        return utc_value.strftime("%Y-%m-%dT%H:%M:%S.000Z")
    except (TypeError, ValueError):
        return None


def _load_video_catalog() -> list[dict[str, str]]:
    if not VIDEO_LIBRARY_CATALOG.exists():
        return []
    try:
        entries = json.loads(VIDEO_LIBRARY_CATALOG.read_text(encoding="utf-8-sig"))
    except json.JSONDecodeError:
        return []
    return [entry for entry in entries if isinstance(entry, dict)]


def _load_audio_catalog() -> list[dict[str, str]]:
    entries: list[dict[str, str]] = []
    if AUDIO_LIBRARY_CATALOG.exists():
        try:
            loaded = json.loads(AUDIO_LIBRARY_CATALOG.read_text(encoding="utf-8-sig"))
            if isinstance(loaded, list):
                for entry in loaded:
                    if not isinstance(entry, dict):
                        continue
                    filename = str(entry.get("filename") or "").strip()
                    if not filename:
                        continue
                    candidate = AUDIO_LIBRARY_DIR / filename
                    if not candidate.exists() or not candidate.is_file():
                        continue
                    entries.append(
                        {
                            "code": str(entry.get("code") or "").strip(),
                            "title": str(entry.get("title") or "").strip(),
                            "filename": filename,
                        }
                    )
        except json.JSONDecodeError:
            entries = []
    if entries:
        return entries

    # Fallback: auto-index supported audio files when catalog is missing.
    discovered: list[dict[str, str]] = []
    for candidate in AUDIO_LIBRARY_DIR.rglob("*"):
        if (
            not candidate.is_file()
            or candidate.suffix.lower() not in {".mp3", ".wav", ".m4a", ".aac", ".flac", ".ogg"}
        ):
            continue
        relative = candidate.relative_to(AUDIO_LIBRARY_DIR).as_posix()
        stem = candidate.stem.replace("_", " ").replace("-", " ").strip()
        discovered.append({"code": candidate.stem, "title": stem, "filename": relative})
    discovered.sort(key=lambda item: str(item.get("filename") or "").lower())
    return discovered


def _load_branding_config() -> dict:
    if not BRANDING_FILE.exists():
        return {}
    try:
        loaded = json.loads(BRANDING_FILE.read_text(encoding="utf-8-sig"))
    except json.JSONDecodeError:
        return {}
    return loaded if isinstance(loaded, dict) else {"packs": loaded}


def _load_branding_packs() -> list[dict]:
    loaded = _load_branding_config()
    if isinstance(loaded, dict):
        candidate_packs = loaded.get("packs")
        if isinstance(candidate_packs, list):
            return [p for p in candidate_packs if isinstance(p, dict)]
        return []
    if isinstance(loaded, list):
        return [p for p in loaded if isinstance(p, dict)]
    return []


def _default_branding_pack() -> dict | None:
    config = _load_branding_config()
    packs = _load_branding_packs()
    if not packs:
        return None
    default_ref = str(config.get("default_pack") or "").strip().lower() if isinstance(config, dict) else ""
    if default_ref:
        for pack in packs:
            keys = {
                str(pack.get("id") or "").strip().lower(),
                str(pack.get("code") or "").strip().lower(),
                str(pack.get("name") or "").strip().lower(),
            }
            if default_ref in keys:
                return pack
    for pack in packs:
        if str(pack.get("id") or "").strip().lower() == "default_branding":
            return pack
    return packs[0]


def _branding_pack_by_reference(reference: object) -> dict | None:
    ref = str(reference or "").strip().lower()
    if not ref:
        return None
    for pack in _load_branding_packs():
        keys = {
            str(pack.get("id") or "").strip().lower(),
            str(pack.get("code") or "").strip().lower(),
            str(pack.get("name") or "").strip().lower(),
            str(pack.get("title") or "").strip().lower(),
        }
        if ref in keys:
            return pack
    return None


def _normalize_category(value: object) -> str:
    normalized = re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())
    return normalized


def _load_video_index() -> list[dict]:
    entries: list = []
    if VIDEO_LIBRARY_INDEX.exists():
        try:
            loaded = json.loads(VIDEO_LIBRARY_INDEX.read_text(encoding="utf-8-sig"))
            if isinstance(loaded, list):
                entries = loaded
        except json.JSONDecodeError:
            entries = []

    normalized_entries = []
    for entry in entries:
        if not isinstance(entry, dict):
            continue
        filename = str(entry.get("filename") or "").strip()
        if not filename:
            continue
        candidate = VIDEO_LIBRARY_DIR / filename
        if not candidate.exists():
            continue
        tags = entry.get("tags")
        if not isinstance(tags, list):
            tags = []
        normalized_entries.append(
            {
                "code": str(entry.get("code") or "").strip(),
                "title": str(entry.get("title") or "").strip(),
                "filename": filename,
                "path": candidate,
                "category": _normalize_category(entry.get("category") or ""),
                "tags": [str(tag).strip().lower() for tag in tags if str(tag).strip()],
            }
        )
    if normalized_entries:
        return normalized_entries

    # Fallback: derive a minimal index from catalog so context switching
    # still works even when video_index.json is missing/invalid.
    fallback_entries = []
    for catalog_entry in _load_video_catalog():
        filename = str(catalog_entry.get("filename") or "").strip()
        if not filename:
            continue
        candidate = VIDEO_LIBRARY_DIR / filename
        if not candidate.exists():
            continue
        category = _normalize_category(Path(filename).parts[0] if "/" in filename or "\\" in filename else "")
        if not category:
            category = _normalize_category(catalog_entry.get("code") or "")
        title = str(catalog_entry.get("title") or "").strip()
        stem_tokens = _tokenize_scene_text(Path(filename).stem.replace("_", " "))
        fallback_entries.append(
            {
                "code": str(catalog_entry.get("code") or "").strip(),
                "title": title,
                "filename": filename,
                "path": candidate,
                "category": category,
                "tags": sorted(stem_tokens) if stem_tokens else ([category] if category else []),
            }
        )
    return fallback_entries


def _available_context_categories() -> list[str]:
    categories = {
        str(entry.get("category") or "").strip().lower()
        for entry in _load_video_index()
        if str(entry.get("category") or "").strip()
    }
    return sorted(category for category in categories if category)


def _fallback_context_video_path(category_hint: str | None = None) -> Path | None:
    normalized_hint = _normalize_category(category_hint or "")
    index_entries = _load_video_index()
    if index_entries:
        if normalized_hint:
            for entry in index_entries:
                if entry.get("category") == normalized_hint:
                    candidate = entry.get("path")
                    if isinstance(candidate, Path) and candidate.exists():
                        return candidate
        first_path = index_entries[0].get("path")
        if isinstance(first_path, Path) and first_path.exists():
            return first_path
    for entry in _load_video_catalog():
        filename = str(entry.get("filename") or "").strip()
        if not filename:
            continue
        candidate = VIDEO_LIBRARY_DIR / filename
        if candidate.exists():
            return candidate
    return None


def _tokenize_scene_text(text: str) -> set[str]:
    tokens = re.findall(r"[a-z0-9']+", str(text or "").lower())
    stop_words = {
        "the",
        "and",
        "for",
        "that",
        "this",
        "with",
        "from",
        "your",
        "you",
        "are",
        "was",
        "were",
        "into",
        "about",
        "have",
        "has",
        "had",
        "not",
        "but",
        "they",
        "them",
        "their",
    }
    return {token for token in tokens if len(token) >= 3 and token not in stop_words}


def _split_script_into_scenes(text: str, max_scenes: int = 6) -> list[str]:
    normalized_text = str(text or "").strip()
    line_chunks = [line.strip() for line in normalized_text.splitlines() if line.strip()]
    if len(line_chunks) > 1:
        chunks = line_chunks
    else:
        chunks = [
            chunk.strip()
            for chunk in re.split(r"(?<=[.!?])\s+|\n+", normalized_text)
            if chunk and chunk.strip()
        ]
    if not chunks:
        return [normalized_text]
    max_scenes = max(1, min(12, int(max_scenes)))
    if len(chunks) == 1 and max_scenes > 1:
        # Fallback for scripts that are one long line with weak punctuation:
        # split by word windows so requested context scene count is honored.
        words = [word for word in re.split(r"\s+", normalized_text) if word]
        if len(words) >= max_scenes:
            per_scene = max(1, (len(words) + max_scenes - 1) // max_scenes)
            forced_chunks = []
            for idx in range(0, len(words), per_scene):
                forced_chunks.append(" ".join(words[idx : idx + per_scene]).strip())
            chunks = [chunk for chunk in forced_chunks if chunk]
        elif len(words) > 1:
            # For very short one-liners, keep one or two words per scene.
            chunks = words
        elif len(normalized_text) >= 8:
            # Last fallback for tiny text with one token.
            char_slices = []
            width = max(1, len(normalized_text) // max_scenes)
            for idx in range(0, len(normalized_text), width):
                char_slices.append(normalized_text[idx : idx + width].strip())
            chunks = [chunk for chunk in char_slices if chunk]
    if len(chunks) <= max_scenes:
        return chunks
    merged = []
    group_size = (len(chunks) + max_scenes - 1) // max_scenes
    for idx in range(0, len(chunks), group_size):
        merged.append(" ".join(chunks[idx : idx + group_size]).strip())
    return [scene for scene in merged if scene]


def _scene_important_tokens(scene_text: str) -> list[str]:
    words = re.findall(r"[a-z0-9']+", str(scene_text or "").lower())
    stop_words = {
        "the",
        "and",
        "for",
        "that",
        "this",
        "with",
        "from",
        "your",
        "you",
        "are",
        "was",
        "were",
        "into",
        "about",
        "have",
        "has",
        "had",
        "not",
        "but",
        "they",
        "them",
        "their",
        "will",
        "just",
        "then",
        "than",
    }
    return [w for w in words if len(w) >= 3 and w not in stop_words]


def _candidate_semantic_score(
    *,
    scene_text: str,
    scene_tokens: set[str],
    important_tokens: list[str],
    candidate_tokens: set[str],
    category_match: bool,
) -> int:
    if not candidate_tokens:
        return 0
    base_overlap = len(scene_tokens.intersection(candidate_tokens))
    important_overlap = sum(1 for token in important_tokens if token in candidate_tokens)

    phrase_bonus = 0
    scene_lower = scene_text.lower()
    for phrase in re.findall(r"[a-z0-9']+\s+[a-z0-9']+", scene_lower):
        a, b = phrase.split(" ", 1)
        if a in candidate_tokens and b in candidate_tokens:
            phrase_bonus += 1

    score = (base_overlap * 3) + (important_overlap * 2) + phrase_bonus
    if category_match:
        score += 2
    return score


def _scene_durations_by_text_weight(scenes: list[str], total_duration: float) -> list[float]:
    if not scenes:
        return []
    if total_duration <= 0:
        return [1.0 for _ in scenes]
    weights = [max(1, len(_tokenize_scene_text(scene))) for scene in scenes]
    total_weight = float(sum(weights))
    durations = [max(0.8, total_duration * (weight / total_weight)) for weight in weights]
    scaled_total = sum(durations)
    if scaled_total > 0:
        scale = total_duration / scaled_total
        durations = [max(0.6, duration * scale) for duration in durations]
    delta = total_duration - sum(durations)
    durations[-1] = max(0.6, durations[-1] + delta)
    return durations


def _row_video_strategy(row: dict) -> str:
    strategy = str(
        row.get("video_strategy")
        or row.get("video_mode")
        or row.get("video_selection_mode")
        or "single"
    ).strip().lower()
    aliases = {
        "default": "single",
        "single_video": "single",
        "single_library": "single",
        "single": "single",
        "context": "context_switch",
        "contextual": "context_switch",
        "context_switching": "context_switch",
        "smart": "context_switch",
        "auto_context": "context_switch",
    }
    return aliases.get(strategy, strategy)


def _scene_candidate_score(
    scene_tokens: set[str], candidate_tokens: set[str], category_boost: bool
) -> int:
    if not candidate_tokens:
        return 0
    overlap = len(scene_tokens.intersection(candidate_tokens))
    return overlap + (2 if category_boost else 0)


def _scene_category_scores(
    scene_text: str,
    scene_tokens: set[str],
    available_categories: set[str],
) -> dict[str, int]:
    normalized_scene = str(scene_text or "").lower()
    scores: dict[str, int] = {}
    for category in available_categories:
        rule = CONTEXT_CATEGORY_RULES.get(category)
        if not rule:
            continue
        token_matches = len(scene_tokens.intersection(set(rule.get("tokens", set()))))
        phrase_hits = sum(
            1 for phrase in rule.get("phrases", set()) if phrase in normalized_scene
        )
        scores[category] = token_matches + (phrase_hits * 3)
    return scores


def _infer_scene_category(
    scene_text: str,
    scene_tokens: set[str],
    available_categories: set[str],
    category_hint: str,
) -> str | None:
    scores = _scene_category_scores(scene_text, scene_tokens, available_categories)
    if category_hint in available_categories:
        scores[category_hint] = scores.get(category_hint, 0) + 1
    if not scores:
        return category_hint if category_hint in available_categories else None
    best_category, best_score = max(scores.items(), key=lambda item: item[1])
    if best_score <= 0:
        return category_hint if category_hint in available_categories else None
    return best_category


def _select_context_clip_paths(
    scenes: list[str],
    row: dict,
    fallback_video_path: Path,
) -> tuple[list[Path], list[str]]:
    if not scenes:
        return [fallback_video_path], ["fallback"]

    category_hint = _normalize_category(
        row.get("category_hint") or row.get("content_category") or row.get("video_category")
    )
    category_lock = _to_bool(
        row.get("context_lock_category"),
        default=bool(category_hint),
    )
    index_entries = _load_video_index()
    if not index_entries:
        return [fallback_video_path for _ in scenes], ["fallback" for _ in scenes]

    candidates = []
    category_buckets: dict[str, list[dict]] = {}
    for entry in index_entries:
        tags = set(_tokenize_scene_text(" ".join(entry.get("tags", []))))
        title_tokens = _tokenize_scene_text(str(entry.get("title") or ""))
        category = str(entry.get("category") or "")
        filename_tokens = _tokenize_scene_text(str(entry.get("filename") or ""))
        category_tokens = _tokenize_scene_text(category.replace("_", " "))
        code = str(entry.get("code") or "")
        candidate = {
            "path": entry["path"],
            "category": category,
            "tokens": tags.union(title_tokens).union(category_tokens).union(filename_tokens),
            "code": code,
        }
        candidates.append(
            candidate
        )
        if category:
            category_buckets.setdefault(category, []).append(candidate)
    if not candidates:
        return [fallback_video_path for _ in scenes], ["fallback" for _ in scenes]
    for category in category_buckets:
        category_buckets[category] = sorted(
            category_buckets[category], key=lambda candidate: candidate["code"]
        )

    selected_paths: list[Path] = []
    selected_categories: list[str] = []
    last_selected: Path | None = None
    used_paths: set[str] = set()
    category_cursor: dict[str, int] = {}
    available_categories = set(category_buckets.keys())
    locked_category: str | None = None
    if category_lock and category_hint in available_categories:
        locked_category = category_hint
    for scene in scenes:
        scene_tokens = _tokenize_scene_text(scene)
        important_tokens = _scene_important_tokens(scene)
        if locked_category:
            target_category = locked_category
        else:
            target_category = _infer_scene_category(
                scene, scene_tokens, available_categories, category_hint
            )

        ranked_pool = category_buckets.get(target_category or "", [])
        if ranked_pool:
            ranked_local = sorted(
                ranked_pool,
                key=lambda candidate: (
                    _candidate_semantic_score(
                        scene_text=scene,
                        scene_tokens=scene_tokens,
                        important_tokens=important_tokens,
                        candidate_tokens=candidate["tokens"],
                        category_match=True,
                    ),
                    candidate["code"],
                ),
                reverse=True,
            )
            picked_local = None
            for candidate in ranked_local:
                key = str(candidate["path"])
                if candidate["path"] == last_selected and len(ranked_local) > 1:
                    continue
                if key in used_paths and len(ranked_local) > 1:
                    continue
                picked_local = candidate
                break
            if picked_local is None:
                cursor = category_cursor.get(target_category or "", 0)
                picked_local = ranked_local[cursor % len(ranked_local)]
                category_cursor[target_category or ""] = cursor + 1
            selected_paths.append(picked_local["path"])
            selected_categories.append(target_category or "fallback")
            last_selected = picked_local["path"]
            used_paths.add(str(picked_local["path"]))
            continue

        ranked = sorted(
            candidates,
            key=lambda candidate: (
                _candidate_semantic_score(
                    scene_text=scene,
                    scene_tokens=scene_tokens,
                    important_tokens=important_tokens,
                    candidate_tokens=candidate["tokens"],
                    category_match=bool(
                        target_category and candidate["category"] == target_category
                    ),
                ),
                1 if target_category and candidate["category"] == target_category else 0,
                candidate["code"],
            ),
            reverse=True,
        )
        chosen_path = fallback_video_path
        chosen_category = "fallback"
        for option in ranked:
            option_path = option["path"]
            option_key = str(option_path)
            if last_selected is not None and option_path == last_selected and len(ranked) > 1:
                continue
            if option_key in used_paths and len(ranked) > 1:
                continue
            chosen_path = option_path
            chosen_category = option["category"] or "fallback"
            break
        selected_paths.append(chosen_path)
        selected_categories.append(chosen_category)
        last_selected = chosen_path
        used_paths.add(str(chosen_path))
    # Force diversity when scene_count > 1: if all selections collapsed to one clip,
    # rotate through same-category pool first, then global candidates.
    if len(scenes) > 1 and len({str(path) for path in selected_paths}) == 1:
        diversity_pool: list[dict] = []
        if category_hint and category_hint in category_buckets:
            diversity_pool = category_buckets[category_hint]
        if not diversity_pool:
            diversity_pool = sorted(candidates, key=lambda candidate: candidate["code"])
        if len(diversity_pool) > 1:
            diversified_paths: list[Path] = []
            diversified_categories: list[str] = []
            for idx in range(len(scenes)):
                pick = diversity_pool[idx % len(diversity_pool)]
                diversified_paths.append(pick["path"])
                diversified_categories.append(pick["category"] or "fallback")
            selected_paths = diversified_paths
            selected_categories = diversified_categories

    # Additional dedup pass: replace repeated selections where possible, while
    # preserving per-scene category intent.
    if len(scenes) > 1 and selected_paths:
        seen_paths: set[str] = set()
        all_pool = sorted(candidates, key=lambda candidate: candidate["code"])
        for idx, current_path in enumerate(selected_paths):
            current_key = str(current_path)
            if current_key not in seen_paths:
                seen_paths.add(current_key)
                continue

            target_category = selected_categories[idx] if idx < len(selected_categories) else ""
            replacement_pool = category_buckets.get(target_category, [])
            if not replacement_pool:
                replacement_pool = all_pool

            replacement = None
            for candidate in replacement_pool:
                key = str(candidate["path"])
                if key in seen_paths:
                    continue
                replacement = candidate
                break

            if replacement is not None:
                selected_paths[idx] = replacement["path"]
                if idx < len(selected_categories):
                    selected_categories[idx] = replacement["category"] or target_category or "fallback"
                seen_paths.add(str(replacement["path"]))
            else:
                seen_paths.add(current_key)

    return selected_paths, selected_categories


def _build_context_background_video(
    *,
    job_id: str,
    text: str,
    total_duration: float,
    output_mode: str,
    row: dict,
    fallback_video_path: Path,
) -> tuple[Path, int, list[str], list[str]]:
    max_scenes_value = row.get("context_scene_count")
    try:
        max_scenes = max(1, min(12, int(max_scenes_value)))
    except (TypeError, ValueError):
        max_scenes = 6
    scenes = _split_script_into_scenes(text, max_scenes=max_scenes)
    scene_durations = _scene_durations_by_text_weight(scenes, total_duration)
    planned_crossfade = 0.0
    if len(scene_durations) >= 2:
        # xfade overlaps adjacent segments, so stitched duration becomes shorter.
        # Pre-compensate by adding the overlap budget to the last segment.
        min_scene_duration = min(scene_durations)
        planned_crossfade = min(0.35, max(0.0, min_scene_duration - 0.15))
        overlap_budget = planned_crossfade * (len(scene_durations) - 1)
        if overlap_budget > 0:
            scene_durations[-1] = scene_durations[-1] + overlap_budget
    scene_clip_paths, scene_categories = _select_context_clip_paths(
        scenes, row, fallback_video_path
    )
    if len(scene_clip_paths) > 1 and len({str(path) for path in scene_clip_paths}) == 1:
        # Final guard: force unique clip paths before segment rendering.
        hint = _normalize_category(
            row.get("category_hint") or row.get("content_category") or row.get("video_category")
        )
        index_entries = _load_video_index()
        pool = []
        for entry in index_entries:
            p = entry.get("path")
            if not isinstance(p, Path) or not p.exists():
                continue
            category = str(entry.get("category") or "")
            if hint and category != hint:
                continue
            pool.append(p)
        if len(pool) < 2:
            pool = [entry.get("path") for entry in index_entries if isinstance(entry.get("path"), Path) and entry.get("path").exists()]
        unique_pool = []
        seen = set()
        for p in pool:
            key = str(p)
            if key in seen:
                continue
            seen.add(key)
            unique_pool.append(p)
        if len(unique_pool) > 1:
            scene_clip_paths = [unique_pool[idx % len(unique_pool)] for idx in range(len(scene_clip_paths))]

    segment_paths = []
    for idx, (clip_path, segment_duration) in enumerate(
        zip(scene_clip_paths, scene_durations), start=1
    ):
        segment_output = OUTPUT_DIR / f"{job_id}_ctxseg_{idx:03}.mp4"
        frame_filter = _build_smart_framing_filter(clip_path, output_mode)
        segment_filter = f"{frame_filter},fps=30,settb=AVTB,format=yuv420p"
        clip_duration = _video_duration_seconds(clip_path)
        max_seek = max(0.0, clip_duration - max(0.6, float(segment_duration)) - 0.1)
        seek_offset = 0.0
        if max_seek > 0:
            # Vary start frame to avoid visual repetition even when a clip is reused.
            seek_offset = min(max_seek, (idx * 1.37) % max_seek)
        segment_command = [
            "ffmpeg",
            "-y",
            "-ss",
            f"{seek_offset:.2f}",
            "-stream_loop",
            "-1",
            "-i",
            str(clip_path),
            "-t",
            f"{max(0.6, float(segment_duration)):.2f}",
            "-vf",
            segment_filter,
            "-an",
            "-c:v",
            "libx264",
            "-pix_fmt",
            "yuv420p",
            str(segment_output),
        ]
        _run_ffmpeg(segment_command)
        segment_paths.append(segment_output)

    stitched_output = OUTPUT_DIR / f"{job_id}_context_bg.mp4"
    # Smooth transitions: crossfade scene segments into a single continuous background.
    # Fallback to concat when there are too few/too-short segments for crossfade.
    def _concat_segments() -> None:
        concat_list_path = OUTPUT_DIR / f"{job_id}_ctx_concat.txt"
        concat_lines = []
        for segment_path in segment_paths:
            concat_file = str(segment_path.resolve()).replace("\\", "/").replace("'", "'\\''")
            concat_lines.append(f"file '{concat_file}'")
        concat_list_path.write_text("\n".join(concat_lines) + "\n", encoding="utf-8")
        concat_command = [
            "ffmpeg",
            "-y",
            "-f",
            "concat",
            "-safe",
            "0",
            "-i",
            str(concat_list_path),
            "-an",
            "-c:v",
            "libx264",
            "-pix_fmt",
            "yuv420p",
            str(stitched_output),
        ]
        _run_ffmpeg(concat_command)

    if len(segment_paths) >= 2:
        durations = [_video_duration_seconds(path) for path in segment_paths]
        min_duration = min(durations) if durations else 0.0
        crossfade = min(planned_crossfade, min_duration - 0.15)
        crossfade = max(0.0, crossfade)
        if crossfade >= 0.12:
            xfade_inputs: list[str] = []
            xfade_filter_parts: list[str] = []
            for segment_path in segment_paths:
                xfade_inputs.extend(["-i", str(segment_path)])

            offset = max(0.0, durations[0] - crossfade)
            xfade_filter_parts.append(
                f"[0:v][1:v]xfade=transition=fade:duration={crossfade:.2f}:offset={offset:.2f}[v1]"
            )
            cumulative_duration = durations[0]
            for idx in range(2, len(segment_paths)):
                cumulative_duration += durations[idx - 1] - crossfade
                offset = max(0.0, cumulative_duration - crossfade)
                xfade_filter_parts.append(
                    f"[v{idx-1}][{idx}:v]xfade=transition=fade:duration={crossfade:.2f}:offset={offset:.2f}[v{idx}]"
                )

            last_label = f"[v{len(segment_paths)-1}]"
            xfade_command = [
                "ffmpeg",
                "-y",
                *xfade_inputs,
                "-filter_complex",
                ";".join(xfade_filter_parts),
                "-map",
                last_label,
                "-an",
                "-c:v",
                "libx264",
                "-pix_fmt",
                "yuv420p",
                str(stitched_output),
            ]
            try:
                _run_ffmpeg(xfade_command)
            except RuntimeError:
                _concat_segments()
        else:
            _concat_segments()
    else:
        _concat_segments()
    unique_codes = len({path.name for path in scene_clip_paths})
    scene_clip_names = [path.name for path in scene_clip_paths]
    return stitched_output, unique_codes, scene_categories, scene_clip_names


def _library_video_path(code: str) -> Path | None:
    for entry in _load_video_catalog():
        if entry.get("code") == code:
            filename = entry.get("filename")
            if filename:
                candidate = VIDEO_LIBRARY_DIR / filename
                return candidate if candidate.exists() else None
    return None


def _library_video_by_reference(reference: str) -> tuple[Path, dict] | None:
    ref = str(reference or "").strip().lower()
    if not ref:
        return None
    for entry in _load_video_catalog():
        filename = str(entry.get("filename") or "").strip()
        if not filename:
            continue
        candidate = VIDEO_LIBRARY_DIR / filename
        if not candidate.exists():
            continue
        code = str(entry.get("code") or "").strip().lower()
        title = str(entry.get("title") or "").strip().lower()
        base = Path(filename).stem.lower()
        full = filename.lower()
        if ref in {code, title, base, full}:
            return candidate, entry
    return None


def _library_audio_by_reference(reference: str) -> tuple[Path, dict] | None:
    ref = str(reference or "").strip().lower()
    if not ref:
        return None
    for entry in _load_audio_catalog():
        filename = str(entry.get("filename") or "").strip()
        if not filename:
            continue
        candidate = AUDIO_LIBRARY_DIR / filename
        if not candidate.exists() or not candidate.is_file():
            continue
        code = str(entry.get("code") or "").strip().lower()
        title = str(entry.get("title") or "").strip().lower()
        base = Path(filename).stem.lower()
        full = filename.lower()
        if ref in {code, title, base, full}:
            return candidate, entry
    return None


def _load_script_from_excel(excel_path: Path, video_code: str) -> str | None:
    workbook = load_workbook(filename=excel_path, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
    try:
        code_index = headers.index("video_code")
    except ValueError:
        code_index = -1
    text_index = headers.index("script_text") if "script_text" in headers else -1
    file_index = headers.index("script_file") if "script_file" in headers else -1
    if code_index == -1:
        return None

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        current_code = str(row[code_index]).strip() if row[code_index] is not None else ""
        if current_code != video_code:
            continue
        if text_index != -1 and row[text_index]:
            return str(row[text_index]).strip()
        if file_index != -1 and row[file_index]:
            script_name = _safe_filename(str(row[file_index]).strip())
            script_path = SCRIPTS_DIR / script_name
            if script_path.exists():
                return script_path.read_text(encoding="utf-8")
    return None


def _wav_duration_seconds(wav_path: Path) -> float:
    with wave.open(str(wav_path), "rb") as wav_file:
        frames = wav_file.getnframes()
        rate = wav_file.getframerate()
        return frames / float(rate)


def _audio_duration_seconds(audio_path: Path) -> float:
    probe_command = [
        "ffprobe",
        "-v",
        "error",
        "-show_entries",
        "format=duration",
        "-of",
        "default=noprint_wrappers=1:nokey=1",
        str(audio_path),
    ]
    process = subprocess.run(probe_command, capture_output=True, text=True)
    if process.returncode == 0:
        try:
            return float(process.stdout.strip())
        except ValueError:
            pass

    with wave.open(str(audio_path), "rb") as wav_file:
        frames = wav_file.getnframes()
        rate = wav_file.getframerate()
        return frames / float(rate)


def _text_to_speech(
    text: str,
    output_path: Path,
    rate: int,
    voice_style: str,
    voice_gender: str,
) -> tuple[str, str]:
    style = VOICE_STYLE_PRESETS.get(voice_style, VOICE_STYLE_PRESETS["professional"])
    profile = _voice_profile(voice_style, rate)
    style_models = VOICE_MODEL_MATRIX.get(voice_style, VOICE_MODEL_MATRIX["professional"])
    model_name = style_models.get(voice_gender, style_models["male"])
    model_path = PIPER_VOICES_DIR / model_name

    piper_executable = shutil.which("piper")
    if not piper_executable:
        candidate = Path(sys.executable).resolve().parent / (
            "piper.exe" if os.name == "nt" else "piper"
        )
        if candidate.exists():
            piper_executable = str(candidate)

    if piper_executable and model_path.exists():
        input_text_path = output_path.with_suffix(".txt")
        input_text_path.write_text(text, encoding="utf-8")
        piper_command = [
            piper_executable,
            "--model",
            str(model_path),
            "--input_file",
            str(input_text_path),
            "--output_file",
            str(output_path),
            "--length_scale",
            f"{profile['length_scale']:.2f}",
            "--noise_scale",
            f"{profile['noise_scale']:.2f}",
            "--noise_w_scale",
            f"{profile['noise_w_scale']:.2f}",
            "--volume",
            f"{profile['volume']:.2f}",
        ]
        process = subprocess.run(piper_command, capture_output=True, text=True)
        input_text_path.unlink(missing_ok=True)
        if process.returncode == 0 and output_path.exists():
            _post_process_voice(output_path, profile)
            return "piper", model_name

    adjusted_rate = max(120, min(260, int(rate * (1 + (style["rate_adjust"] / 100)))))
    engine = pyttsx3.init()
    engine.setProperty("rate", adjusted_rate)
    engine.save_to_file(text, str(output_path))
    engine.runAndWait()
    return "pyttsx3", "pyttsx3-default"


def _format_ass_time(seconds: float) -> str:
    hours = int(seconds // 3600)
    minutes = int((seconds % 3600) // 60)
    seconds_remainder = seconds % 60
    return f"{hours}:{minutes:02d}:{seconds_remainder:05.2f}"


def _hex_to_ass_color(hex_color: str, alpha: str = "00") -> str:
    value = hex_color.strip().lstrip("#")
    if len(value) != 6:
        value = "FFFFFF"
    rr = value[0:2]
    gg = value[2:4]
    bb = value[4:6]
    return f"&H{alpha}{bb}{gg}{rr}".upper()


def _build_ass_style_line(
    text_color: str,
    bg_color: str,
    bold: bool,
    italic: bool,
    alignment: int = 2,
    font_size: int = 48,
    font_name: str = "Arial",
    border_style: int = 3,
    outline: int = 1,
    shadow: int = 0,
    margin_v: int = 80,
) -> str:
    primary = _hex_to_ass_color(text_color, alpha="00")
    back = _hex_to_ass_color(bg_color, alpha="00")

    bold_value = -1 if bold else 0
    italic_value = -1 if italic else 0

    return (
        f"Style: Default,{font_name},"
        f"{font_size},{primary},&H000000FF,{back},{back},"
        f"{bold_value},{italic_value},0,0,100,100,0,0,{border_style},{outline},{shadow},{alignment},60,60,{margin_v},1"
    )


def _subtitle_segments_from_text(text: str, max_words_per_segment: int = 9) -> list[str]:
    normalized = " ".join(str(text or "").replace("\n", " ").split()).strip()
    if not normalized:
        return []

    sentence_parts = [
        part.strip()
        for part in re.findall(r"[^.!?]+[.!?]?", normalized)
        if part and part.strip()
    ]
    if not sentence_parts:
        sentence_parts = [normalized]

    segments: list[str] = []
    for sentence in sentence_parts:
        clause_parts = [
            clause.strip()
            for clause in re.split(r"(?<=[,;:])\s+|\s+(?:and|but|because|so|while)\s+", sentence)
            if clause and clause.strip()
        ]
        if not clause_parts:
            clause_parts = [sentence]
        for clause in clause_parts:
            words = clause.split()
            if not words:
                continue
            if len(words) <= max_words_per_segment:
                segments.append(clause.strip())
                continue

            has_terminal_punct = bool(re.search(r"[.!?]\s*$", clause))
            clean_clause = clause.rstrip(".!? ").strip()
            clean_words = clean_clause.split()
            for idx in range(0, len(clean_words), max_words_per_segment):
                chunk_words = clean_words[idx : idx + max_words_per_segment]
                if not chunk_words:
                    continue
                chunk_text = " ".join(chunk_words)
                if has_terminal_punct and idx + max_words_per_segment >= len(clean_words):
                    chunk_text = chunk_text + "."
                segments.append(chunk_text.strip())
    return segments


def _estimate_word_syllables(word: str) -> int:
    token = re.sub(r"[^a-z]", "", str(word or "").lower())
    if not token:
        return 1
    groups = re.findall(r"[aeiouy]+", token)
    count = len(groups)
    if token.endswith("e") and count > 1:
        count -= 1
    return max(1, count)


def _tokenize_subtitle_words(line: str) -> list[tuple[str, str]]:
    tokens = [token for token in re.split(r"\s+", str(line or "").strip()) if token]
    pairs: list[tuple[str, str]] = []
    for token in tokens:
        normalized = re.sub(r"(^[^a-z0-9']+|[^a-z0-9']+$)", "", token.lower())
        pairs.append((token, normalized))
    return pairs


def _subtitle_keyword_set(text: str, max_keywords: int = 14) -> set[str]:
    counts: dict[str, int] = {}
    for token in re.findall(r"[a-z0-9']+", str(text or "").lower()):
        if len(token) < 4:
            continue
        if token in {
            "that",
            "this",
            "with",
            "from",
            "your",
            "have",
            "will",
            "just",
            "they",
            "them",
            "then",
            "when",
            "what",
            "where",
            "there",
            "would",
            "could",
            "about",
            "into",
            "while",
            "because",
            "video",
        }:
            continue
        counts[token] = counts.get(token, 0) + 1
    ranked = sorted(counts.items(), key=lambda item: (-item[1], -len(item[0]), item[0]))
    return {token for token, _ in ranked[:max_keywords]}


def _ass_primary_color_override(hex_color: str) -> str:
    value = str(hex_color or "").strip().lstrip("#")
    if len(value) != 6:
        value = "FFD35A"
    rr = value[0:2]
    gg = value[2:4]
    bb = value[4:6]
    return f"&H{bb}{gg}{rr}&".upper()


def _beat_sync_word_durations_cs(
    words: list[tuple[str, str]],
    segment_duration: float,
) -> list[int]:
    if not words:
        return []
    total_cs = max(1, int(round(max(0.45, segment_duration) * 100)))
    weights: list[float] = []
    for raw_word, normalized_word in words:
        base = float(_estimate_word_syllables(normalized_word or raw_word))
        # Tiny punctuation pause so emphatic words can pop on-beat.
        if re.search(r"[,.!?;:]$", raw_word):
            base += 0.35
        weights.append(max(0.5, base))
    total_weight = sum(weights) or float(len(words))
    durations = [max(1, int(round(total_cs * (w / total_weight)))) for w in weights]
    delta = total_cs - sum(durations)
    durations[-1] = max(1, durations[-1] + delta)
    return durations


def _build_beat_sync_dialogue(
    line: str,
    segment_duration: float,
    *,
    line_index: int,
    keyword_set: set[str],
    emphasis_color: str,
) -> str:
    words = _tokenize_subtitle_words(line)
    if not words:
        return str(line or "").replace("{", "\\{").replace("}", "\\}")
    durations_cs = _beat_sync_word_durations_cs(words, segment_duration)
    parts: list[str] = []
    accent = _ass_primary_color_override(emphasis_color)
    for idx, ((raw_word, normalized_word), word_cs) in enumerate(zip(words, durations_cs)):
        safe_word = raw_word.replace("{", "\\{").replace("}", "\\}")
        is_keyword = normalized_word in keyword_set
        looks_important = (
            len(normalized_word) >= 8
            or re.search(r"\d", normalized_word or "")
            or str(raw_word).isupper()
        )
        is_emphasis = is_keyword or looks_important
        # Rhythm pulse every 2-3 words to feel edited without overwhelming readability.
        has_pulse = ((idx + line_index) % 3 == 0) or ((idx + line_index) % 2 == 0 and is_emphasis)
        if is_emphasis:
            tag = (
                f"{{\\k{word_cs}\\1c{accent}\\bord3"
                "\\t(0,95,\\fscx132\\fscy132)"
                "\\t(95,215,\\fscx100\\fscy100)}"
            )
        elif has_pulse:
            tag = (
                f"{{\\k{word_cs}"
                "\\t(0,80,\\fscx114\\fscy114)"
                "\\t(80,175,\\fscx100\\fscy100)}"
            )
        else:
            tag = f"{{\\k{word_cs}}}"
        separator = " " if idx < len(words) - 1 else ""
        parts.append(f"{tag}{safe_word}{separator}")
    return "".join(parts)


def _subtitle_segment_weights(segments: list[str]) -> list[float]:
    weights: list[float] = []
    for segment in segments:
        words = [word for word in segment.split() if word]
        syllables = sum(_estimate_word_syllables(word) for word in words)
        # Blend syllable complexity with word count to better approximate TTS pacing.
        base = max(1.0, (syllables * 0.72) + (len(words) * 0.45))
        punctuation_bonus = 0.0
        if re.search(r"[.!?]\s*$", segment):
            punctuation_bonus += 1.25
        elif re.search(r"[,;:]\s*$", segment):
            punctuation_bonus += 0.60
        weights.append(base + punctuation_bonus)
    return weights


def _normalize_segment_durations(weights: list[float], total_duration: float) -> list[float]:
    if not weights:
        return []
    if total_duration <= 0:
        return [1.0 / len(weights) for _ in weights]

    total_weight = sum(weights) or float(len(weights))
    durations = [total_duration * (weight / total_weight) for weight in weights]
    count = len(durations)
    min_duration = max(0.14, min(0.36, total_duration / max(1, count * 1.55)))

    if total_duration > (min_duration * count):
        for idx, value in enumerate(durations):
            if value < min_duration:
                durations[idx] = min_duration
        current_total = sum(durations)
        if current_total > total_duration:
            excess = current_total - total_duration
            adjustable = [
                idx for idx, value in enumerate(durations) if value > min_duration + 1e-6
            ]
            while excess > 1e-6 and adjustable:
                per_slot = excess / len(adjustable)
                next_adjustable: list[int] = []
                for idx in adjustable:
                    reducible = durations[idx] - min_duration
                    reduction = min(reducible, per_slot)
                    durations[idx] -= reduction
                    excess -= reduction
                    if durations[idx] > min_duration + 1e-6:
                        next_adjustable.append(idx)
                adjustable = next_adjustable

    delta = total_duration - sum(durations)
    durations[-1] = max(0.06, durations[-1] + delta)
    return durations


def _subtitle_timing_probe_with_tts(
    segments: list[str],
    *,
    tts_rate: int,
    voice_style: str,
    voice_gender: str,
    max_segments: int = 24,
) -> list[float] | None:
    """Measure per-segment speaking time using the same TTS pipeline.

    This improves subtitle/audio sync but can be slower on long scripts.
    """
    if not segments:
        return None
    if len(segments) > max_segments:
        return None

    durations: list[float] = []
    probe_id = uuid.uuid4().hex
    for idx, segment in enumerate(segments, start=1):
        if not segment.strip():
            durations.append(0.08)
            continue
        probe_path = OUTPUT_DIR / f"{probe_id}_subprobe_{idx:03}.wav"
        try:
            _text_to_speech(
                segment,
                probe_path,
                tts_rate,
                voice_style,
                voice_gender,
            )
            segment_duration = _audio_duration_seconds(probe_path)
            if re.search(r"[.!?]\s*$", segment):
                segment_duration += 0.12
            elif re.search(r"[,;:]\s*$", segment):
                segment_duration += 0.06
            durations.append(max(0.06, float(segment_duration)))
        except Exception:
            return None
        finally:
            probe_path.unlink(missing_ok=True)
    return durations


def _build_ass_subtitles(
    text: str,
    total_duration: float,
    output_path: Path,
    text_color: str,
    bg_color: str,
    bold: bool,
    italic: bool,
    alignment: int = 2,
    font_size: int = 48,
    font_name: str = "Arial",
    border_style: int = 3,
    outline: int = 1,
    shadow: int = 0,
    margin_v: int = 80,
    play_res_x: int = 1280,
    play_res_y: int = 720,
    subtitle_template: str = "fade",
    tts_rate: int = 175,
    voice_style: str = "professional",
    voice_gender: str = "male",
    precise_timing: bool = True,
) -> None:
    subtitle_template = _resolve_subtitle_template(subtitle_template)
    lines = _subtitle_segments_from_text(text, max_words_per_segment=9)
    if not lines:
        lines = [str(text or "").strip()]
    measured_durations = None
    if precise_timing:
        measured_durations = _subtitle_timing_probe_with_tts(
            lines,
            tts_rate=tts_rate,
            voice_style=voice_style,
            voice_gender=voice_gender,
        )

    if measured_durations and len(measured_durations) == len(lines):
        weights = [max(0.06, value) for value in measured_durations]
    else:
        weights = _subtitle_segment_weights(lines)
    segment_durations = _normalize_segment_durations(weights, max(0.01, total_duration))

    if subtitle_template == "bold_center":
        alignment = 5
        bold = True
        font_size = max(font_size, 54)
        outline = max(outline, 2)
        shadow = 0
        margin_v = max(42, int(margin_v * 0.65))
    elif subtitle_template == "beat_sync":
        outline = max(outline, 2)
        shadow = max(shadow, 1)
        font_size = max(font_size, 52)
        margin_v = max(42, int(margin_v * 0.72))

    style_line = _build_ass_style_line(
        text_color,
        bg_color,
        bold,
        italic,
        alignment=alignment,
        font_size=font_size,
        font_name=font_name,
        border_style=border_style,
        outline=outline,
        shadow=shadow,
        margin_v=margin_v,
    )

    header = f"""[Script Info]
ScriptType: v4.00+
Collisions: Normal
PlayResX: {play_res_x}
PlayResY: {play_res_y}

[V4+ Styles]
Format: Name, Fontname, Fontsize, PrimaryColour, SecondaryColour, OutlineColour, BackColour, Bold, Italic, Underline, StrikeOut, ScaleX, ScaleY, Spacing, Angle, BorderStyle, Outline, Shadow, Alignment, MarginL, MarginR, MarginV, Encoding
{style_line}

[Events]
Format: Layer, Start, End, Style, Name, MarginL, MarginR, MarginV, Effect, Text
"""

    events = []
    beat_keywords = _subtitle_keyword_set(text) if subtitle_template == "beat_sync" else set()
    start_time = 0.0
    for line_index, (line, segment_duration) in enumerate(zip(lines, segment_durations)):
        end_time = start_time + max(0.08, segment_duration)
        start_stamp = _format_ass_time(start_time)
        end_stamp = _format_ass_time(end_time)
        escaped = line.replace("{", "\\{").replace("}", "\\}")
        dialogue_text = f"{{\\fad(200,200)}}{escaped}"
        if subtitle_template == "bold_center":
            dialogue_text = f"{{\\an5\\b1\\fad(140,160)}}{escaped}"
        elif subtitle_template == "bounce_fade":
            dialogue_text = (
                "{\\fad(120,220)\\t(0,120,\\fscx130\\fscy130)"
                "\\t(120,280,\\fscx100\\fscy100)}"
                f"{escaped}"
            )
        elif subtitle_template == "karaoke_word_by_word":
            words = [w for w in line.split() if w]
            if words:
                total_cs = max(1, int(round(max(0.5, segment_duration) * 100)))
                per_word = max(1, total_cs // len(words))
                parts = []
                consumed = 0
                for idx, word in enumerate(words):
                    if idx == len(words) - 1:
                        word_cs = max(1, total_cs - consumed)
                    else:
                        word_cs = per_word
                        consumed += word_cs
                    safe_word = word.replace("{", "\\{").replace("}", "\\}")
                    separator = " " if idx < len(words) - 1 else ""
                    parts.append(f"{{\\k{word_cs}}}{safe_word}{separator}")
                dialogue_text = "".join(parts)
        elif subtitle_template == "beat_sync":
            dialogue_text = _build_beat_sync_dialogue(
                line,
                segment_duration,
                line_index=line_index,
                keyword_set=beat_keywords,
                emphasis_color="#FFD35A",
            )
        events.append(
            f"Dialogue: 0,{start_stamp},{end_stamp},Default,,0,0,0,,{dialogue_text}"
        )
        start_time = end_time

    output_path.write_text(header + "\n" + "\n".join(events), encoding="utf-8")


def _hex_to_ass_color_batch(hex_str: str | None, default: str = "&H00FFFFFF") -> str:
    """Convert #RRGGBB or #AARRGGBB to ASS &HAABBGGRR."""
    if not hex_str or not isinstance(hex_str, str):
        return default
    hex_str = hex_str.strip().lstrip("#")
    if len(hex_str) == 6:
        r, g, b = hex_str[0:2], hex_str[2:4], hex_str[4:6]
        a = "00"
    elif len(hex_str) == 8:
        a, r, g, b = hex_str[0:2], hex_str[2:4], hex_str[4:6], hex_str[6:8]
    else:
        return default
    return f"&H{a}{b}{g}{r}"


def _placement_to_ass_alignment(placement: str | None) -> int:
    """ASS Alignment: 2=bottom center, 5=middle, 8=top."""
    if not placement:
        return 2
    p = str(placement).strip().lower()
    if p in ("top", "upper"):
        return 8
    if p in ("middle", "center", "mid"):
        return 5
    return 2


def _to_bool(value: object, default: bool = False) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "yes", "y", "on"}


def _resolve_bgm_volume(value: object, default: float = 0.18) -> float:
    if value is None or str(value).strip() == "":
        return default
    try:
        parsed = float(str(value).strip())
    except (TypeError, ValueError):
        return default
    return max(0.0, min(1.0, parsed))


def _resolve_bgm_reference(reference: object) -> Path | None:
    ref = str(reference or "").strip()
    if not ref:
        return None
    audio_match = _library_audio_by_reference(ref)
    if audio_match:
        return audio_match[0]
    candidate = Path(ref)
    if candidate.exists() and candidate.is_file():
        return candidate
    safe_name = _safe_filename(ref)
    if not safe_name:
        return None
    audio_candidate = AUDIO_LIBRARY_DIR / safe_name
    if audio_candidate.exists() and audio_candidate.is_file():
        return audio_candidate
    local_candidate = MUSIC_DIR / safe_name
    if local_candidate.exists() and local_candidate.is_file():
        return local_candidate
    return None


def _row_audio_reference(row: dict) -> str:
    for key in (
        "audio_library",
        "audio_library_code",
        "bgm_library",
        "bgm_library_code",
        "audio_code",
        "audio_name",
        "audio_filename",
        "bgm_file",
        "bgm_track",
        "background_music",
        "music_file",
    ):
        value = row.get(key)
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def _resolve_batch_bgm_path(row: dict) -> Path | None:
    reference = _row_audio_reference(row)
    if not reference:
        return None
    return _resolve_bgm_reference(reference)


def _resolve_batch_asset_path(
    row: dict, keys: tuple[str, ...], default_dir: Path
) -> Path | None:
    for key in keys:
        value = row.get(key)
        if value is None or str(value).strip() == "":
            continue
        reference = str(value).strip()
        candidate = Path(reference)
        if candidate.exists() and candidate.is_file():
            return candidate
        safe_name = _safe_filename(reference)
        if not safe_name:
            continue
        local_candidate = default_dir / safe_name
        if local_candidate.exists() and local_candidate.is_file():
            return local_candidate
    return None


def _row_branding_reference(row: dict) -> str:
    for key in ("branding_pack", "branding_pack_id", "brand_pack", "brand_code"):
        value = row.get(key)
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def _coerce_int(value: object, default: int, min_value: int, max_value: int) -> int:
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        parsed = default
    return max(min_value, min(max_value, parsed))


def _branding_value(
    row: dict, pack: dict | None, keys: tuple[str, ...], default: object = ""
) -> object:
    for key in keys:
        value = row.get(key)
        if value is not None and str(value).strip() != "":
            return value
    if pack is not None:
        for key in keys:
            if key in pack and pack.get(key) is not None and str(pack.get(key)).strip() != "":
                return pack.get(key)
    return default


def _build_bgm_audio_filter_chain(
    *,
    bgm_volume: float,
    bgm_ducking: bool,
    voice_stream_label: str = "1:a",
    bgm_stream_label: str = "2:a",
) -> str:
    volume = max(0.0, min(1.0, bgm_volume))
    # Boost narration without reducing user-selected BGM level.
    if bgm_ducking:
        return (
            f"[{voice_stream_label}]volume=1.30[voice];"
            f"[{bgm_stream_label}]volume={volume:.2f}[bgm];"
            "[bgm][voice]sidechaincompress="
            "threshold=0.035:ratio=10:attack=12:release=220[ducked];"
            "[voice][ducked]amix=inputs=2:weights='1.0 1.0':duration=first:dropout_transition=2[aout]"
        )
    return (
        f"[{voice_stream_label}]volume=1.30[voice];"
        f"[{bgm_stream_label}]volume={volume:.2f}[bgm];"
        "[voice][bgm]amix=inputs=2:weights='1.0 1.0':duration=first:dropout_transition=2[aout]"
    )


def _batch_alignment_value(row: dict, default_alignment: int = 2) -> int:
    # Supports either numeric "alignment" (Create Video style) or textual "placement".
    raw_alignment = row.get("alignment") or row.get("subtitle_alignment")
    if raw_alignment is not None and str(raw_alignment).strip() != "":
        try:
            alignment = int(raw_alignment)
            if alignment in {1, 2, 3, 4, 5, 6, 7, 8, 9}:
                return alignment
        except (TypeError, ValueError):
            pass
    raw_placement = (
        row.get("placement") or row.get("subtitle_placement") or row.get("subtitle_pl")
    )
    if raw_placement is not None and str(raw_placement).strip():
        return _placement_to_ass_alignment(raw_placement)
    return default_alignment


def _build_ass_subtitles_styled(
    text: str,
    total_duration: float,
    output_path: Path,
    *,
    font_color: str | None = None,
    font_background: str | None = None,
    alignment: int = 2,
) -> None:
    """Build ASS subtitles with configurable font color, background, and placement."""
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    if not lines:
        lines = [text.strip()]
    word_counts = [max(1, len(line.split())) for line in lines]
    total_words = sum(word_counts)
    primary = _hex_to_ass_color_batch(font_color, "&H00FFFFFF")
    back = _hex_to_ass_color_batch(font_background, "&H64000000")
    header = f"""[Script Info]
ScriptType: v4.00+
Collisions: Normal
PlayResX: 1280
PlayResY: 720

[V4+ Styles]
Format: Name, Fontname, Fontsize, PrimaryColour, SecondaryColour, OutlineColour, BackColour, Bold, Italic, Underline, StrikeOut, ScaleX, ScaleY, Spacing, Angle, BorderStyle, Outline, Shadow, Alignment, MarginL, MarginR, MarginV, Encoding
Style: Default,Arial,48,{primary},&H000000FF,&H00000000,{back},0,0,0,0,100,100,0,0,1,3,2,{alignment},60,60,80,1

[Events]
Format: Layer, Start, End, Style, Name, MarginL, MarginR, MarginV, Effect, Text
"""
    events = []
    start_time = 0.0
    for line, count in zip(lines, word_counts):
        segment_duration = (count / total_words) * total_duration
        end_time = start_time + max(0.5, segment_duration)
        start_stamp = _format_ass_time(start_time)
        end_stamp = _format_ass_time(end_time)
        escaped = line.replace("{", "\\{").replace("}", "\\}")
        events.append(
            f"Dialogue: 0,{start_stamp},{end_stamp},Default,,0,0,0,,{{\\fad(200,200)}}{escaped}"
        )
        start_time = end_time
    output_path.write_text(header + "\n" + "\n".join(events), encoding="utf-8")


def _text_to_speech_with_voice(
    text: str, output_path: Path, rate: int, voice_id: str | None = None
) -> None:
    """TTS with optional voice selection (pyttsx3 voice id or name)."""
    engine = pyttsx3.init()
    engine.setProperty("rate", rate)
    if voice_id and str(voice_id).strip():
        for v in engine.getProperty("voices"):
            if voice_id.strip() in (v.id, getattr(v, "name", "")):
                engine.setProperty("voice", v.id)
                break
    engine.save_to_file(text, str(output_path))
    engine.runAndWait()


def _normalize_header(h: str) -> str:
    """Normalize Excel column header for lookup: lowercase, spaces to underscores."""
    if not h:
        return ""
    return str(h).strip().lower().replace(" ", "_")


def _row_library_reference(row: dict) -> str:
    """
    Resolve batch library selector precedence.
    Supported columns: video_code, library_video_code, library_video,
    library_video_name, library_filename, source_video.
    """
    for key in (
        "video_code",
        "library_video_code",
        "library_video",
        "library_video_name",
        "library_filename",
        "source_video",
    ):
        value = row.get(key)
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def _parse_batch_excel(excel_path: Path) -> list[dict]:
    """
    Parse Excel into list of row dicts. Each row must have a library selector
    and script.
    Library selector columns: video_code, library_video_code, library_video,
    library_video_name, library_filename, source_video.
    Script columns: video_script, script_text, or script_file.
    """
    workbook = load_workbook(filename=excel_path, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [_normalize_header(cell.value) for cell in sheet[1]]
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        row_dict = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                v = val
                if v is not None and not isinstance(v, (int, float)):
                    v = str(v).strip()
                row_dict[headers[i]] = v
        # Resolve script: video_script, script_text, or script_file
        script = row_dict.get("video_script") or row_dict.get("script_text")
        if not script and row_dict.get("script_file"):
            fn = _safe_filename(str(row_dict.get("script_file", "")))
            if fn:
                sp = SCRIPTS_DIR / fn
                if sp.exists():
                    row_dict["_script_text"] = sp.read_text(encoding="utf-8")
                else:
                    row_dict["_script_text"] = None
            else:
                row_dict["_script_text"] = None
        else:
            row_dict["_script_text"] = script
        library_ref = _row_library_reference(row_dict)
        if library_ref:
            row_dict["_library_ref"] = library_ref
        if row_dict.get("_library_ref") and row_dict.get("_script_text"):
            rows.append(row_dict)
    return rows


def _process_one_batch_row(row: dict) -> dict:
    """
    Process one Excel row: resolve library video, TTS, ASS, FFmpeg.
    Returns dict with job_id, output_url, video_name, video_description, video_tags.
    Raises on error (e.g. library code not found).
    """
    job_id = uuid.uuid4().hex
    library_ref = str(row.get("_library_ref") or _row_library_reference(row)).strip()
    match = _library_video_by_reference(library_ref)
    if not match:
        raise ValueError(f"Library video not found: {library_ref}")
    video_path, video_entry = match
    source_video_path = video_path
    video_code = str(video_entry.get("code") or library_ref).strip()
    text = row.get("_script_text") or ""
    tts_rate = 175
    try:
        r = row.get("tts_rate")
        if r is not None:
            tts_rate = max(120, min(240, int(r)))
    except (TypeError, ValueError):
        pass
    voice_style = str(row.get("voice_style") or "professional").strip().lower()
    if voice_style not in VOICE_STYLE_PRESETS:
        voice_style = "professional"
    voice_gender = str(row.get("voice_gender") or "male").strip().lower()
    if voice_gender not in {"male", "female"}:
        voice_gender = "male"

    voice_path = OUTPUT_DIR / f"{job_id}_voice.wav"
    # Backward compatibility: use explicit pyttsx3 voice id/name if provided.
    if row.get("voice_type"):
        _text_to_speech_with_voice(
            text,
            voice_path,
            tts_rate,
            voice_id=str(row.get("voice_type")).strip(),
        )
        tts_engine = "pyttsx3"
        voice_model = "pyttsx3-selected-voice"
    else:
        tts_engine, voice_model = _text_to_speech(
            text, voice_path, tts_rate, voice_style, voice_gender
        )

    duration = _wav_duration_seconds(voice_path)
    subtitles_path = OUTPUT_DIR / f"{job_id}.ass"
    subtitle_preset = str(
        row.get("subtitle_preset")
        or row.get("subtitle_style")
        or "classic"
    ).strip().lower()
    subtitle_template = _resolve_subtitle_template(
        row.get("subtitle_template")
        or row.get("subtitle_animation")
        or row.get("ass_template")
        or "fade"
    )
    output_mode = _resolve_output_mode(
        row.get("output_mode") or row.get("aspect_mode") or row.get("format_mode")
    )
    video_strategy = _row_video_strategy(row)
    video_source_mode = "single_library"
    context_clip_count = 1
    context_scene_categories: list[str] = []
    context_scene_clips: list[str] = []
    context_error = None
    if video_strategy == "context_switch":
        try:
            stitched_path, clip_count, scene_categories, scene_clips = _build_context_background_video(
                job_id=job_id,
                text=text,
                total_duration=duration,
                output_mode=output_mode,
                row=row,
                fallback_video_path=source_video_path,
            )
            video_path = stitched_path
            video_source_mode = "context_switch"
            context_clip_count = max(1, clip_count)
            context_scene_categories = scene_categories
            context_scene_clips = scene_clips
        except Exception as exc:
            # Do not fail the row; fallback to single library video.
            video_path = source_video_path
            context_error = str(exc)
    target_w, target_h = _output_mode_resolution(output_mode)
    preset_style = _resolve_subtitle_style_preset(subtitle_preset)
    text_color = str(
        row.get("text_color")
        or row.get("subtitle_text_color")
        or row.get("subtitle_tc")
        or row.get("font_color")
        or preset_style["text_color"]
    ).strip()
    bg_color = str(
        row.get("bg_color")
        or row.get("subtitle_bg_color")
        or row.get("subtitle_bg")
        or row.get("font_background")
        or preset_style["bg_color"]
    ).strip()
    bold_value = _to_bool(
        row.get("bold")
        if row.get("bold") is not None
        else row.get("subtitle_bold", row.get("subtitle_b")),
        default=bool(preset_style["bold"]),
    )
    italic_value = _to_bool(
        row.get("italic")
        if row.get("italic") is not None
        else row.get("subtitle_italic", row.get("subtitle_it")),
        default=bool(preset_style["italic"]),
    )
    alignment_value = _batch_alignment_value(
        row, default_alignment=int(preset_style["alignment"])
    )
    margin_v_value = _subtitle_margin_for_mode(
        int(preset_style["margin_v"]), output_mode, target_h
    )
    _build_ass_subtitles(
        text,
        duration,
        subtitles_path,
        text_color=text_color,
        bg_color=bg_color,
        bold=bold_value,
        italic=italic_value,
        alignment=alignment_value,
        font_size=int(preset_style["font_size"]),
        font_name=str(preset_style["font_name"]),
        border_style=int(preset_style["border_style"]),
        outline=int(preset_style["outline"]),
        shadow=int(preset_style["shadow"]),
        margin_v=margin_v_value,
        play_res_x=target_w,
        play_res_y=target_h,
        subtitle_template=subtitle_template,
        tts_rate=tts_rate,
        voice_style=voice_style,
        voice_gender=voice_gender,
        precise_timing=True,
    )
    output_path = OUTPUT_DIR / f"{job_id}_final.mp4"
    frame_filter = _build_smart_framing_filter(video_path, output_mode)
    subtitles_filter = (
        f"subtitles=filename='{_ffmpeg_subtitles_path(subtitles_path)}':charenc=UTF-8"
    )
    video_filter_chain = f"{frame_filter},{subtitles_filter}"
    bgm_path = _resolve_batch_bgm_path(row)
    bgm_volume = _resolve_bgm_volume(row.get("bgm_volume") or row.get("music_volume"))
    bgm_ducking = _to_bool(row.get("bgm_ducking"), default=True)
    use_looped_video_input = video_source_mode != "context_switch"
    video_input_args = ["-i", str(video_path)]
    if use_looped_video_input:
        video_input_args = ["-stream_loop", "-1", "-i", str(video_path)]

    if bgm_path:
        audio_filter_chain = _build_bgm_audio_filter_chain(
            bgm_volume=bgm_volume,
            bgm_ducking=bgm_ducking,
            voice_stream_label="1:a",
            bgm_stream_label="2:a",
        )
        ffmpeg_command = [
            "ffmpeg",
            "-y",
            *video_input_args,
            "-i",
            str(voice_path),
            "-stream_loop",
            "-1",
            "-i",
            str(bgm_path),
            "-t",
            f"{duration:.2f}",
            "-vf",
            video_filter_chain,
            "-filter_complex",
            audio_filter_chain,
            "-map",
            "0:v:0",
            "-map",
            "[aout]",
            "-c:v",
            "libx264",
            "-c:a",
            "aac",
            str(output_path),
        ]
    else:
        ffmpeg_command = [
            "ffmpeg",
            "-y",
            *video_input_args,
            "-i",
            str(voice_path),
            "-t",
            f"{duration:.2f}",
            "-vf",
            video_filter_chain,
            "-c:v",
            "libx264",
            "-c:a",
            "aac",
            str(output_path),
        ]
    try:
        _run_ffmpeg(ffmpeg_command)
    except RuntimeError as primary_error:
        # Some FFmpeg builds do not include sidechaincompress. Retry without ducking.
        if bgm_path and bgm_ducking:
            fallback_audio_filter_chain = _build_bgm_audio_filter_chain(
                bgm_volume=bgm_volume,
                bgm_ducking=False,
                voice_stream_label="1:a",
                bgm_stream_label="2:a",
            )
            fallback_command = [
                "ffmpeg",
                "-y",
                *video_input_args,
                "-i",
                str(voice_path),
                "-stream_loop",
                "-1",
                "-i",
                str(bgm_path),
                "-t",
                f"{duration:.2f}",
                "-vf",
                video_filter_chain,
                "-filter_complex",
                fallback_audio_filter_chain,
                "-map",
                "0:v:0",
                "-map",
                "[aout]",
                "-c:v",
                "libx264",
                "-c:a",
                "aac",
                str(output_path),
            ]
            try:
                _run_ffmpeg(fallback_command)
                ffmpeg_command = fallback_command
                bgm_ducking = False
            except RuntimeError as fallback_error:
                raise RuntimeError(str(fallback_error))
        else:
            raise RuntimeError(str(primary_error))
    output_video_name = row.get("output_video_name")
    if output_video_name is not None and str(output_video_name).strip():
        video_name = str(output_video_name).strip()
    else:
        video_name = row.get("video_name")
        if video_name is not None and str(video_name).strip():
            video_name = str(video_name).strip()
        else:
            video_name = str(video_entry.get("title") or video_code).strip()
    payload = {
        "job_id": job_id,
        "output_url": f"/api/download/{output_path.name}",
        "video_name": video_name,
        "video_description": str(row.get("video_description") or ""),
        "video_tags": str(row.get("video_tags") or ""),
        "library_video_code": video_code,
        "library_video_title": str(video_entry.get("title") or ""),
        "tts_engine": tts_engine,
        "voice_style": voice_style,
        "voice_gender": voice_gender,
        "voice_model": voice_model,
        "subtitle_preset": subtitle_preset,
        "subtitle_template": subtitle_template,
        "output_mode": output_mode,
        "video_source_mode": video_source_mode,
        "video_strategy": video_strategy,
        "context_clip_count": context_clip_count,
        "context_scene_categories": context_scene_categories,
        "context_scene_clips": context_scene_clips,
        "bgm_enabled": bool(bgm_path),
        "audio_library_ref": _row_audio_reference(row) if bgm_path else "",
        "bgm_ducking": bgm_ducking if bgm_path else False,
        "bgm_volume": bgm_volume if bgm_path else 0.0,
    }
    if context_error:
        payload["context_error"] = context_error
    return payload


@app.post("/api/process")
def process_video(
    background_video: UploadFile | None = File(None),
    script_file: UploadFile | None = File(None),
    bgm_file: UploadFile | None = File(None),
    audio_library_ref: str = Form(""),
    library_code: str = Form(""),
    script_excel: UploadFile | None = File(None),
    tts_rate: int = Form(175),
    text_color: str | None = Form(None),
    bg_color: str | None = Form(None),
    bold: str | None = Form(None),
    italic: str | None = Form(None),
    alignment: int | None = Form(None),
    subtitle_preset: str = Form("classic"),
    subtitle_template: str = Form("fade"),
    output_mode: str = Form("youtube"),
    video_strategy: str = Form("single"),
    category_hint: str = Form(""),
    context_scene_count: int = Form(6),
    context_lock_category: str = Form("true"),
    bgm_volume: str | None = Form(None),
    bgm_ducking: str | None = Form(None),
    voice_style: str = Form("professional"),
    voice_gender: str = Form("male"),
) -> JSONResponse:
    preset_style = _resolve_subtitle_style_preset(subtitle_preset)
    text_color = str(text_color or preset_style["text_color"]).strip()
    bg_color = str(bg_color or preset_style["bg_color"]).strip()
    is_bold = _to_bool(bold, default=bool(preset_style["bold"]))
    is_italic = _to_bool(italic, default=bool(preset_style["italic"]))
    alignment_value = int(preset_style["alignment"])
    if alignment is not None:
        try:
            parsed_alignment = int(alignment)
            if parsed_alignment in {1, 2, 3, 4, 5, 6, 7, 8, 9}:
                alignment_value = parsed_alignment
        except (TypeError, ValueError):
            pass

    subtitle_preset = str(subtitle_preset or "classic").strip().lower()
    if subtitle_preset == "default":
        subtitle_preset = "classic"
    if subtitle_preset not in SUBTITLE_STYLE_PRESETS:
        subtitle_preset = "classic"
    subtitle_template = _resolve_subtitle_template(subtitle_template)
    output_mode = _resolve_output_mode(output_mode)
    video_strategy = _row_video_strategy({"video_strategy": video_strategy})
    target_w, target_h = _output_mode_resolution(output_mode)
    voice_style = (voice_style or "professional").strip().lower()
    if voice_style not in VOICE_STYLE_PRESETS:
        voice_style = "professional"
    voice_gender = (voice_gender or "male").strip().lower()
    if voice_gender not in {"male", "female"}:
        voice_gender = "male"

    job_id = uuid.uuid4().hex
    video_path: Path | None = None
    if library_code:
        video_path = _library_video_path(library_code)
        if video_path is None:
            return JSONResponse({"error": "Library video code not found."}, status_code=400)
    elif background_video is not None:
        video_path = UPLOAD_DIR / f"{job_id}_{_safe_filename(background_video.filename)}"
        with video_path.open("wb") as video_buffer:
            shutil.copyfileobj(background_video.file, video_buffer)
    elif video_strategy == "context_switch":
        video_path = _fallback_context_video_path(category_hint)
        if video_path is None:
            return JSONResponse(
                {
                    "error": "No library videos available for context switching. Add videos to data/video_library and catalog/index first."
                },
                status_code=400,
            )
    else:
        return JSONResponse({"error": "Provide a background video or select a library code."}, status_code=400)

    text = None
    if script_excel is not None and library_code:
        excel_path = UPLOAD_DIR / f"{job_id}_{_safe_filename(script_excel.filename)}"
        with excel_path.open("wb") as excel_buffer:
            shutil.copyfileobj(script_excel.file, excel_buffer)
        text = _load_script_from_excel(excel_path, library_code)
        if text is None:
            return JSONResponse({"error": "No script found in Excel for this video code."}, status_code=400)

    if text is None:
        if script_file is None:
            return JSONResponse({"error": "Provide a script text file or Excel mapping."}, status_code=400)
        script_path = UPLOAD_DIR / f"{job_id}_{_safe_filename(script_file.filename)}"
        with script_path.open("wb") as script_buffer:
            shutil.copyfileobj(script_file.file, script_buffer)
        text = script_path.read_text(encoding="utf-8")

    voice_path = OUTPUT_DIR / f"{job_id}_voice.wav"
    tts_engine, voice_model = _text_to_speech(
        text, voice_path, tts_rate, voice_style, voice_gender
    )
    bgm_volume_value = _resolve_bgm_volume(bgm_volume, default=0.18)
    bgm_ducking_value = _to_bool(bgm_ducking, default=True)
    bgm_path: Path | None = None
    if bgm_file is not None and bgm_file.filename and str(bgm_file.filename).strip():
        bgm_path = UPLOAD_DIR / f"{job_id}_{_safe_filename(bgm_file.filename)}"
        with bgm_path.open("wb") as bgm_buffer:
            shutil.copyfileobj(bgm_file.file, bgm_buffer)
    elif audio_library_ref and str(audio_library_ref).strip():
        bgm_path = _resolve_bgm_reference(audio_library_ref)
        if bgm_path is None:
            return JSONResponse(
                {"error": f"Audio library track not found: {str(audio_library_ref).strip()}"},
                status_code=400,
            )

    duration = _audio_duration_seconds(voice_path)
    source_video_path = video_path
    if library_code:
        video_source_mode = "single_library"
    elif background_video is not None:
        video_source_mode = "single_upload"
    else:
        video_source_mode = "single_context_fallback"
    context_clip_count = 1
    context_scene_categories: list[str] = []
    context_scene_clips: list[str] = []
    context_error = None
    if video_strategy == "context_switch":
        try:
            stitched_path, clip_count, scene_categories, scene_clips = _build_context_background_video(
                job_id=job_id,
                text=text,
                total_duration=duration,
                output_mode=output_mode,
                row={
                    "category_hint": category_hint,
                    "context_scene_count": context_scene_count,
                    "context_lock_category": context_lock_category,
                },
                fallback_video_path=source_video_path,
            )
            video_path = stitched_path
            video_source_mode = "context_switch"
            context_clip_count = max(1, clip_count)
            context_scene_categories = scene_categories
            context_scene_clips = scene_clips
        except Exception as exc:
            video_path = source_video_path
            context_error = str(exc)

    subtitles_path = OUTPUT_DIR / f"{job_id}.ass"

    _build_ass_subtitles(
        text,
        duration,
        subtitles_path,
        text_color=text_color,
        bg_color=bg_color,
        bold=is_bold,
        italic=is_italic,
        alignment=alignment_value,
        font_size=int(preset_style["font_size"]),
        font_name=str(preset_style["font_name"]),
        border_style=int(preset_style["border_style"]),
        outline=int(preset_style["outline"]),
        shadow=int(preset_style["shadow"]),
        margin_v=_subtitle_margin_for_mode(
            int(preset_style["margin_v"]), output_mode, target_h
        ),
        play_res_x=target_w,
        play_res_y=target_h,
        subtitle_template=subtitle_template,
        tts_rate=tts_rate,
        voice_style=voice_style,
        voice_gender=voice_gender,
        precise_timing=True,
    )

    output_path = OUTPUT_DIR / f"{job_id}_final.mp4"

    frame_filter = _build_smart_framing_filter(video_path, output_mode)
    subtitles_filter = (
        f"subtitles=filename='{_ffmpeg_subtitles_path(subtitles_path)}':charenc=UTF-8"
    )
    video_filter_chain = f"{frame_filter},{subtitles_filter}"
    use_looped_video_input = video_source_mode != "context_switch"
    video_input_args = ["-i", str(video_path)]
    if use_looped_video_input:
        video_input_args = ["-stream_loop", "-1", "-i", str(video_path)]
    if bgm_path:
        audio_filter_chain = _build_bgm_audio_filter_chain(
            bgm_volume=bgm_volume_value,
            bgm_ducking=bgm_ducking_value,
            voice_stream_label="1:a",
            bgm_stream_label="2:a",
        )
        ffmpeg_command = [
            "ffmpeg",
            "-y",
            *video_input_args,
            "-i",
            str(voice_path),
            "-stream_loop",
            "-1",
            "-i",
            str(bgm_path),
            "-t",
            f"{duration:.2f}",
            "-vf",
            video_filter_chain,
            "-filter_complex",
            audio_filter_chain,
            "-map",
            "0:v:0",
            "-map",
            "[aout]",
            "-c:v",
            "libx264",
            "-c:a",
            "aac",
            str(output_path),
        ]
    else:
        ffmpeg_command = [
            "ffmpeg",
            "-y",
            *video_input_args,
            "-i",
            str(voice_path),
            "-t",
            f"{duration:.2f}",
            "-vf",
            video_filter_chain,
            "-c:v",
            "libx264",
            "-c:a",
            "aac",
            str(output_path),
        ]

    try:
        _run_ffmpeg(ffmpeg_command)
    except RuntimeError as primary_error:
        # Some FFmpeg builds do not include sidechaincompress. Retry without ducking.
        if bgm_path and bgm_ducking_value:
            fallback_audio_filter_chain = _build_bgm_audio_filter_chain(
                bgm_volume=bgm_volume_value,
                bgm_ducking=False,
                voice_stream_label="1:a",
                bgm_stream_label="2:a",
            )
            fallback_command = [
                "ffmpeg",
                "-y",
                *video_input_args,
                "-i",
                str(voice_path),
                "-stream_loop",
                "-1",
                "-i",
                str(bgm_path),
                "-t",
                f"{duration:.2f}",
                "-vf",
                video_filter_chain,
                "-filter_complex",
                fallback_audio_filter_chain,
                "-map",
                "0:v:0",
                "-map",
                "[aout]",
                "-c:v",
                "libx264",
                "-c:a",
                "aac",
                str(output_path),
            ]
            try:
                _run_ffmpeg(fallback_command)
                ffmpeg_command = fallback_command
                bgm_ducking_value = False
            except RuntimeError as fallback_error:
                return JSONResponse({"error": str(fallback_error)}, status_code=500)
        else:
            return JSONResponse({"error": str(primary_error)}, status_code=500)

    payload = {
        "job_id": job_id,
        "output_url": f"/api/download/{output_path.name}",
        "ffmpeg_command": " ".join(ffmpeg_command),
        "tts_engine": tts_engine,
        "voice_style": voice_style,
        "voice_gender": voice_gender,
        "voice_model": voice_model,
        "subtitle_preset": subtitle_preset,
        "subtitle_template": subtitle_template,
        "output_mode": output_mode,
        "video_strategy": video_strategy,
        "video_source_mode": video_source_mode,
        "context_clip_count": context_clip_count,
        "context_scene_categories": context_scene_categories,
        "context_scene_clips": context_scene_clips,
        "bgm_enabled": bool(bgm_path),
        "audio_library_ref": str(audio_library_ref or "").strip() if bgm_path else "",
        "bgm_ducking": bgm_ducking_value if bgm_path else False,
        "bgm_volume": bgm_volume_value if bgm_path else 0.0,
    }
    if context_error:
        payload["context_error"] = context_error
    return JSONResponse(payload)


@app.get("/api/batch/schema")
def batch_schema() -> JSONResponse:
    """Return expected Excel columns for batch upload (for frontend)."""
    return JSONResponse(
        {
            "required": [
                {
                    "key": "video_code",
                    "description": "Primary library selector. Alternatives: library_video_code, library_video_name, library_filename, library_video, source_video.",
                },
                {
                    "key": "video_script",
                    "description": "Script text for TTS and subtitles. Alternative: script_text or script_file (filename in data/scripts).",
                },
            ],
            "optional": [
                {"key": "video_name", "description": "Output title for generated video metadata."},
                {"key": "output_video_name", "description": "Explicit output title override (preferred)."},
                {"key": "library_video_code", "description": "Alternative library selector by catalog code."},
                {"key": "library_video_name", "description": "Alternative library selector by catalog title."},
                {"key": "library_filename", "description": "Alternative library selector by catalog filename."},
                {"key": "library_video", "description": "Alternative generic library selector (code/title/filename)."},
                {"key": "source_video", "description": "Alternative generic library selector (code/title/filename)."},
                {"key": "video_description", "description": "Video description (e.g. for YouTube)."},
                {"key": "video_tags", "description": "Comma-separated tags."},
                {"key": "publish_at", "description": "Optional YouTube schedule time. Supports ISO format; converted to UTC for YouTube."},
                {"key": "visibility", "description": "public, private, or unlisted. If set to private, video stays private even when publish_at is filled."},
                {"key": "output_mode", "description": "youtube (16:9), shorts/reels (9:16), or square (1:1)."},
                {"key": "video_strategy", "description": "single (default) or context_switch for scene-based background switching."},
                {"key": "category_hint", "description": "Optional category hint for context_switch (e.g. motivation, finance, politics)."},
                {"key": "context_scene_count", "description": "Optional max scene segments for context_switch (1-12, default 6)."},
                {"key": "voice_style", "description": "professional, casual, narrator, energetic, calm, dramatic."},
                {"key": "voice_gender", "description": "male or female."},
                {"key": "voice_type", "description": "Optional pyttsx3 voice id/name override."},
                {"key": "thumbnail_mode", "description": "Optional: auto, manual, or none."},
                {"key": "thumbnail_file", "description": "Optional when thumbnail_mode=manual. Filename in data/thumbnails (or absolute path)."},
                {"key": "branding_pack", "description": "Optional reusable branding pack id/code/name from data/branding_packs.json."},
                {"key": "logo_file", "description": "Optional logo image filename in data/logos (or absolute path)."},
                {"key": "logo_position", "description": "Optional: top-left, top-right, bottom-left, bottom-right, center."},
                {"key": "logo_scale_percent", "description": "Optional logo size percent (5-40), default 15."},
                {"key": "logo_animated", "description": "Optional true/false. If true, logo fades in."},
                {"key": "intro_text", "description": "Optional intro branding text overlay."},
                {"key": "intro_duration_sec", "description": "Optional intro overlay duration (1-15)."},
                {"key": "subscribe_cta_text", "description": "Optional CTA text, e.g. Subscribe for more."},
                {"key": "subscribe_cta_duration_sec", "description": "Optional CTA duration (2-20)."},
                {"key": "subscribe_cta_from_end_sec", "description": "Optional CTA start offset from end (2-120)."},
                {"key": "end_screen_blocks", "description": "Optional true/false to draw reusable end-screen placeholders."},
                {"key": "end_screen_duration_sec", "description": "Optional end-screen block duration (2-20)."},
                {"key": "outro_text", "description": "Optional outro branding text overlay."},
                {"key": "outro_duration_sec", "description": "Optional outro overlay duration (1-20)."},
                {"key": "end_credits_text", "description": "Optional outro/end credits text."},
                {"key": "end_credits_duration_sec", "description": "Optional end credits duration (2-30), default 5."},
                {"key": "audio_library", "description": "Optional audio library selector (code/title/filename from data/audio_library/catalog.json)."},
                {"key": "audio_library_code", "description": "Optional alias for audio_library (catalog code)."},
                {"key": "audio_name", "description": "Optional alias for audio_library (catalog title or filename stem)."},
                {"key": "bgm_file", "description": "Optional music filename or path. Supports data/audio_library (catalog/fallback) and data/music."},
                {"key": "bgm_volume", "description": "Optional music volume 0.00-1.00 (default 0.18)."},
                {"key": "bgm_ducking", "description": "Optional true/false to auto-lower music during speech (default true)."},
                {"key": "text_color", "description": "Subtitle text color, e.g. #FFFFFF (alias: font_color)."},
                {"key": "bg_color", "description": "Subtitle background color, e.g. #000000 (alias: font_background)."},
                {"key": "bold", "description": "true/false."},
                {"key": "italic", "description": "true/false."},
                {"key": "subtitle_preset", "description": "classic, viral, reels, or cinematic."},
                {"key": "subtitle_template", "description": "Optional: fade, bold_center, karaoke_word_by_word, bounce_fade, or beat_sync."},
                {"key": "alignment", "description": "ASS alignment 1-9 (Create Video style)."},
                {"key": "placement", "description": "Alias for alignment: top, middle, or bottom."},
                {"key": "tts_rate", "description": "Speech rate (120-240), default 175."},
            ],
        }
    )


@app.get("/api/batch/template")
def batch_template() -> StreamingResponse:
    """Download an Excel template for batch processing."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "batch_videos"
    headers = [
        "video_code",
        "video_script",
        "script_file",
        "output_video_name",
        "video_description",
        "video_tags",
        "publish_at",
        "visibility",
        "output_mode",
        "video_strategy",
        "category_hint",
        "context_scene_count",
        "voice_style",
        "voice_gender",
        "thumbnail_mode",
        "thumbnail_file",
        "branding_pack",
        "logo_file",
        "logo_position",
        "logo_scale_percent",
        "logo_animated",
        "intro_text",
        "intro_duration_sec",
        "subscribe_cta_text",
        "subscribe_cta_duration_sec",
        "subscribe_cta_from_end_sec",
        "end_screen_blocks",
        "end_screen_duration_sec",
        "outro_text",
        "outro_duration_sec",
        "end_credits_text",
        "end_credits_duration_sec",
        "audio_library",
        "bgm_file",
        "bgm_volume",
        "bgm_ducking",
        "subtitle_preset",
        "subtitle_template",
        "subtitle_text_color",
        "subtitle_bg_color",
        "subtitle_bold",
        "subtitle_italic",
        "subtitle_placement",
        "tts_rate",
    ]
    sample = [
        "sample_loop",
        "",
        "story1.txt",
        "My Batch Video",
        "Generated from batch template.",
        "story,automation",
        "2026-03-03T10:00",
        "private",
        "youtube",
        "single",
        "",
        "6",
        "professional",
        "male",
        "auto",
        "",
        "default_branding",
        "",
        "top-right",
        "15",
        "true",
        "Welcome to AutoStoryTube",
        "3",
        "Subscribe for more",
        "5",
        "12",
        "true",
        "8",
        "Thanks for watching",
        "6",
        "",
        "5",
        "sample_music",
        "sample_music.mp3",
        "0.18",
        "true",
        "classic",
        "fade",
        "#FFFFFF",
        "#000000",
        "false",
        "false",
        "bottom",
        "175",
    ]
    sheet.append(headers)
    sheet.append(sample)

    subtitle_template_col = headers.index("subtitle_template") + 1
    subtitle_template_col_letter = chr(64 + subtitle_template_col)
    subtitle_template_validation = DataValidation(
        type="list",
        formula1='"fade,bold_center,karaoke_word_by_word,bounce_fade,beat_sync"',
        allow_blank=True,
    )
    subtitle_template_validation.error = "Use one of: fade, bold_center, karaoke_word_by_word, bounce_fade, beat_sync."
    subtitle_template_validation.errorTitle = "Invalid subtitle_template"
    subtitle_template_validation.prompt = "Optional subtitle animation preset."
    subtitle_template_validation.promptTitle = "subtitle_template"
    sheet.add_data_validation(subtitle_template_validation)
    subtitle_template_validation.add(f"{subtitle_template_col_letter}2:{subtitle_template_col_letter}5000")

    video_strategy_col = headers.index("video_strategy") + 1
    video_strategy_col_letter = chr(64 + video_strategy_col)
    video_strategy_validation = DataValidation(
        type="list",
        formula1='"single,context_switch"',
        allow_blank=True,
    )
    video_strategy_validation.error = "Use single or context_switch."
    video_strategy_validation.errorTitle = "Invalid video_strategy"
    video_strategy_validation.prompt = "single = one library clip; context_switch = scene-based clip switching."
    video_strategy_validation.promptTitle = "video_strategy"
    sheet.add_data_validation(video_strategy_validation)
    video_strategy_validation.add(f"{video_strategy_col_letter}2:{video_strategy_col_letter}5000")

    thumbnail_mode_col = headers.index("thumbnail_mode") + 1
    thumbnail_mode_col_letter = chr(64 + thumbnail_mode_col)
    thumbnail_mode_validation = DataValidation(
        type="list",
        formula1='"auto,manual,none"',
        allow_blank=True,
    )
    thumbnail_mode_validation.error = "Use auto, manual, or none."
    thumbnail_mode_validation.errorTitle = "Invalid thumbnail_mode"
    thumbnail_mode_validation.prompt = "Optional: auto, manual, or none."
    thumbnail_mode_validation.promptTitle = "thumbnail_mode"
    sheet.add_data_validation(thumbnail_mode_validation)
    thumbnail_mode_validation.add(f"{thumbnail_mode_col_letter}2:{thumbnail_mode_col_letter}5000")

    bgm_ducking_col = headers.index("bgm_ducking") + 1
    bgm_ducking_col_letter = chr(64 + bgm_ducking_col)
    bgm_ducking_validation = DataValidation(
        type="list",
        formula1='"true,false"',
        allow_blank=True,
    )
    bgm_ducking_validation.error = "Use true or false."
    bgm_ducking_validation.errorTitle = "Invalid bgm_ducking"
    bgm_ducking_validation.prompt = "Optional: auto-lower music during speech."
    bgm_ducking_validation.promptTitle = "bgm_ducking"
    sheet.add_data_validation(bgm_ducking_validation)
    bgm_ducking_validation.add(f"{bgm_ducking_col_letter}2:{bgm_ducking_col_letter}5000")

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"batch_template_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/api/process-batch")
def process_batch(excel_file: UploadFile = File(...)) -> JSONResponse:
    """
    Accept an Excel file with one row per video.
    Required: one library selector column (video_code or aliases) and one script column
    (video_script/script_text/script_file).
    Optional: output titles, description/tags, voice/style columns, and subtitle style columns.
    Returns { "jobs": [ { job_id, output_url, video_name, video_description, video_tags } | { error, video_code } ] }.
    """
    if not excel_file.filename or not str(excel_file.filename).lower().endswith(
        (".xlsx", ".xls")
    ):
        return JSONResponse(
            {"error": "Please upload an Excel file (.xlsx or .xls)."},
            status_code=400,
        )
    upload_id = uuid.uuid4().hex
    excel_path = UPLOAD_DIR / f"{upload_id}_{_safe_filename(excel_file.filename)}"
    with excel_path.open("wb") as f:
        shutil.copyfileobj(excel_file.file, f)
    rows = _parse_batch_excel(excel_path)
    if not rows:
        return JSONResponse(
            {
                "error": "No valid rows found. Each row must have a library selector (video_code or aliases) and video_script (or script_text or script_file)."
            },
            status_code=400,
        )
    jobs = []
    for row in rows:
        try:
            jobs.append(_process_one_batch_row(row))
        except Exception as e:
            jobs.append(
                {
                    "error": str(e),
                    "video_code": str(row.get("video_code", "")).strip(),
                    "library_ref": str(row.get("_library_ref") or _row_library_reference(row)).strip(),
                }
            )
    return JSONResponse({"jobs": jobs})


@app.post("/api/process-batch-youtube")
def process_batch_youtube(excel_file: UploadFile = File(...)) -> JSONResponse:
    """
    Accept an Excel file with one row per video, generate each video, then upload to YouTube.
    Supports optional row-wise scheduling via publish_at.
    """
    creds = _load_credentials()
    if not creds:
        return JSONResponse({"error": "YouTube not authorized."}, status_code=401)
    if not excel_file.filename or not str(excel_file.filename).lower().endswith(
        (".xlsx", ".xls")
    ):
        return JSONResponse(
            {"error": "Please upload an Excel file (.xlsx or .xls)."},
            status_code=400,
        )

    upload_id = uuid.uuid4().hex
    excel_path = UPLOAD_DIR / f"{upload_id}_{_safe_filename(excel_file.filename)}"
    with excel_path.open("wb") as f:
        shutil.copyfileobj(excel_file.file, f)
    rows = _parse_batch_excel(excel_path)
    if not rows:
        return JSONResponse(
            {
                "error": "No valid rows found. Each row must have a library selector (video_code or aliases) and video_script (or script_text or script_file)."
            },
            status_code=400,
        )

    service = build("youtube", "v3", credentials=creds)
    jobs = []
    for idx, row in enumerate(rows, start=1):
        try:
            generated = _process_one_batch_row(row)
            output_url = str(generated.get("output_url") or "").strip()
            output_name = Path(output_url).name
            upload_video_path = OUTPUT_DIR / output_name
            if not upload_video_path.exists():
                raise FileNotFoundError(f"Generated output not found: {output_name}")
            working_video_path = upload_video_path
            logo_applied = False
            logo_error = None
            branding_applied = False
            branding_error = None
            branding_pack_used = ""
            end_credits_applied = False
            end_credits_error = None

            branding_pack = _branding_pack_by_reference(_row_branding_reference(row))
            if branding_pack is None:
                branding_pack = _default_branding_pack()
            if branding_pack:
                branding_pack_used = str(
                    branding_pack.get("id")
                    or branding_pack.get("code")
                    or branding_pack.get("name")
                    or ""
                ).strip()

            logo_path = _resolve_batch_asset_path(
                row,
                ("logo_file", "logo_path", "batch_logo_file"),
                LOGOS_DIR,
            )
            if logo_path is None and branding_pack is not None:
                logo_reference = _branding_value(
                    row, branding_pack, ("logo_file", "logo_path"), ""
                )
                logo_path = _resolve_batch_asset_path(
                    {"logo_file": logo_reference},
                    ("logo_file",),
                    LOGOS_DIR,
                )
            if logo_path:
                raw_logo_position = str(
                    _branding_value(row, branding_pack, ("logo_position",), "top-right")
                ).strip().lower()
                logo_position = (
                    raw_logo_position
                    if raw_logo_position in {"top-left", "top-right", "bottom-left", "bottom-right", "center"}
                    else "top-right"
                )
                logo_scale_percent = _coerce_int(
                    _branding_value(row, branding_pack, ("logo_scale_percent",), 15),
                    default=15,
                    min_value=5,
                    max_value=40,
                )
                logo_animated = _to_bool(
                    _branding_value(row, branding_pack, ("logo_animated",), False),
                    default=False,
                )
                logo_output_path = OUTPUT_DIR / f"{generated.get('job_id')}_batch_logo.mp4"
                try:
                    _apply_logo_overlay_to_video(
                        video_path=working_video_path,
                        logo_path=logo_path,
                        output_path=logo_output_path,
                        logo_position=logo_position,
                        logo_scale_percent=logo_scale_percent,
                        logo_animated=logo_animated,
                    )
                    working_video_path = logo_output_path
                    logo_applied = True
                except Exception as logo_exc:
                    logo_error = str(logo_exc)

            intro_text = str(
                _branding_value(row, branding_pack, ("intro_text", "intro_title"), "")
            ).strip()
            intro_duration_sec = _coerce_int(
                _branding_value(row, branding_pack, ("intro_duration_sec",), 0),
                default=0,
                min_value=0,
                max_value=15,
            )
            outro_text = str(
                _branding_value(row, branding_pack, ("outro_text", "outro_title"), "")
            ).strip()
            outro_duration_sec = _coerce_int(
                _branding_value(row, branding_pack, ("outro_duration_sec",), 0),
                default=0,
                min_value=0,
                max_value=20,
            )
            subscribe_cta_text = str(
                _branding_value(
                    row,
                    branding_pack,
                    ("subscribe_cta_text", "cta_text", "subscribe_text"),
                    "",
                )
            ).strip()
            subscribe_cta_duration_sec = _coerce_int(
                _branding_value(
                    row,
                    branding_pack,
                    ("subscribe_cta_duration_sec", "cta_duration_sec"),
                    5,
                ),
                default=5,
                min_value=2,
                max_value=20,
            )
            subscribe_cta_from_end_sec = _coerce_int(
                _branding_value(
                    row,
                    branding_pack,
                    ("subscribe_cta_from_end_sec", "cta_from_end_sec"),
                    12,
                ),
                default=12,
                min_value=2,
                max_value=120,
            )
            end_screen_blocks = _to_bool(
                _branding_value(row, branding_pack, ("end_screen_blocks",), False),
                default=False,
            )
            end_screen_duration_sec = _coerce_int(
                _branding_value(row, branding_pack, ("end_screen_duration_sec",), 8),
                default=8,
                min_value=2,
                max_value=20,
            )
            if (
                intro_text
                or outro_text
                or subscribe_cta_text
                or end_screen_blocks
            ):
                branding_output_path = OUTPUT_DIR / f"{generated.get('job_id')}_batch_branding.mp4"
                try:
                    _apply_branding_overlays_to_video(
                        video_path=working_video_path,
                        output_path=branding_output_path,
                        intro_text=intro_text,
                        intro_duration_sec=intro_duration_sec,
                        outro_text=outro_text,
                        outro_duration_sec=outro_duration_sec,
                        subscribe_cta_text=subscribe_cta_text,
                        subscribe_cta_duration_sec=subscribe_cta_duration_sec,
                        subscribe_cta_from_end_sec=subscribe_cta_from_end_sec,
                        end_screen_blocks=end_screen_blocks,
                        end_screen_duration_sec=end_screen_duration_sec,
                    )
                    working_video_path = branding_output_path
                    branding_applied = True
                except Exception as branding_exc:
                    branding_error = str(branding_exc)

            credits_text = str(
                _branding_value(
                    row,
                    branding_pack,
                    ("end_credits_text", "credits_text", "outro_text"),
                    "",
                )
                or ""
            ).strip()
            if credits_text:
                credits_duration = _coerce_int(
                    _branding_value(row, branding_pack, ("end_credits_duration_sec",), 5),
                    default=5,
                    min_value=2,
                    max_value=30,
                )
                credits_output_path = OUTPUT_DIR / f"{generated.get('job_id')}_batch_credits.mp4"
                try:
                    _apply_end_credits_to_video(
                        video_path=working_video_path,
                        output_path=credits_output_path,
                        credits_text=credits_text,
                        credits_duration_sec=credits_duration,
                    )
                    working_video_path = credits_output_path
                    end_credits_applied = True
                except Exception as credits_exc:
                    end_credits_error = str(credits_exc)
            upload_video_path = working_video_path

            title = str(
                row.get("output_video_name")
                or row.get("video_name")
                or generated.get("video_name")
                or f"Batch Video {idx}"
            ).strip()
            description = str(row.get("video_description") or "").strip()
            tags = [tag.strip() for tag in str(row.get("video_tags") or "").split(",") if tag.strip()]

            raw_visibility = str(row.get("visibility") or "private").strip().lower()
            safe_visibility = (
                raw_visibility
                if raw_visibility in ("public", "private", "unlisted")
                else "private"
            )
            publish_at_input = str(row.get("publish_at") or "").strip()
            rfc3339_publish_at = _parse_publish_at(publish_at_input)
            if rfc3339_publish_at and safe_visibility != "private":
                status_dict: dict[str, str] = {
                    "privacyStatus": "private",
                    "publishAt": rfc3339_publish_at,
                }
            else:
                status_dict = {"privacyStatus": safe_visibility}

            request_body = {
                "snippet": {
                    "title": title,
                    "description": description,
                    "tags": tags,
                    "categoryId": "22",
                },
                "status": status_dict,
            }

            upload_request = service.videos().insert(
                part=",".join(request_body.keys()),
                body=request_body,
                media_body=MediaFileUpload(str(upload_video_path), resumable=True),
            )
            response = upload_request.execute()
            uploaded_video_id = str(response.get("id") or "").strip()
            thumbnail_uploaded = False
            thumbnail_error = None
            if uploaded_video_id:
                try:
                    row_output_mode = _resolve_output_mode(
                        row.get("output_mode") or row.get("aspect_mode") or row.get("format_mode")
                    )
                    thumbnail_mode_raw = str(
                        row.get("thumbnail_mode")
                        or row.get("thumbnail_type")
                        or "auto"
                    ).strip().lower()
                    if row_output_mode in {"shorts", "reels", "square"} and thumbnail_mode_raw in {"none", "off", "skip", ""}:
                        thumbnail_mode_raw = "auto"
                    if thumbnail_mode_raw in {"none", "off", "skip"}:
                        thumbnail_uploaded = False
                    elif thumbnail_mode_raw in {"manual", "file", "custom"}:
                        manual_thumbnail_path = _resolve_batch_asset_path(
                            row,
                            ("thumbnail_file", "thumbnail_path", "thumb_file"),
                            THUMBNAILS_DIR,
                        )
                        if manual_thumbnail_path:
                            service.thumbnails().set(
                                videoId=uploaded_video_id,
                                media_body=MediaFileUpload(str(manual_thumbnail_path)),
                            ).execute()
                            thumbnail_uploaded = True
                        else:
                            auto_thumbnail_path = OUTPUT_DIR / f"{uploaded_video_id}_thumb_auto.jpg"
                            _generate_thumbnail_from_video(upload_video_path, auto_thumbnail_path)
                            service.thumbnails().set(
                                videoId=uploaded_video_id,
                                media_body=MediaFileUpload(str(auto_thumbnail_path)),
                            ).execute()
                            thumbnail_uploaded = True
                            thumbnail_error = "thumbnail_file not found; used auto thumbnail fallback."
                    else:
                        auto_thumbnail_path = OUTPUT_DIR / f"{uploaded_video_id}_thumb_auto.jpg"
                        _generate_thumbnail_from_video(upload_video_path, auto_thumbnail_path)
                        service.thumbnails().set(
                            videoId=uploaded_video_id,
                            media_body=MediaFileUpload(str(auto_thumbnail_path)),
                        ).execute()
                        thumbnail_uploaded = True
                except Exception as thumb_exc:
                    thumbnail_error = str(thumb_exc)

            result = {
                **generated,
                "youtube_uploaded": True,
                "youtube_video_id": uploaded_video_id,
                "youtube_video_url": f"https://www.youtube.com/watch?v={uploaded_video_id}",
                "thumbnail_uploaded": thumbnail_uploaded,
                "logo_applied": logo_applied,
                "branding_applied": branding_applied,
                "branding_pack": branding_pack_used,
                "end_credits_applied": end_credits_applied,
            }
            if thumbnail_error:
                result["thumbnail_error"] = thumbnail_error
            if logo_error:
                result["logo_error"] = logo_error
            if branding_error:
                result["branding_error"] = branding_error
            if end_credits_error:
                result["end_credits_error"] = end_credits_error
            if rfc3339_publish_at and safe_visibility != "private":
                result["scheduled_publish_at"] = rfc3339_publish_at
            jobs.append(result)
        except HttpError as exc:
            jobs.append(
                {
                    "error": str(exc),
                    "youtube_uploaded": False,
                    "video_code": str(row.get("video_code", "")).strip(),
                    "library_ref": str(row.get("_library_ref") or _row_library_reference(row)).strip(),
                }
            )
        except Exception as exc:
            jobs.append(
                {
                    "error": str(exc),
                    "youtube_uploaded": False,
                    "video_code": str(row.get("video_code", "")).strip(),
                    "library_ref": str(row.get("_library_ref") or _row_library_reference(row)).strip(),
                }
            )
    return JSONResponse({"jobs": jobs})


@app.get("/api/download/{filename}")
def download_output(filename: str) -> FileResponse:
    # Use only basename to prevent path traversal (e.g. batch output files)
    safe_name = Path(filename).name or filename
    file_path = OUTPUT_DIR / safe_name
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(file_path, media_type="video/mp4", filename=safe_name)


def _load_credentials() -> Credentials | None:
    if not TOKEN_PATH.exists():
        return None
    creds = Credentials.from_authorized_user_file(str(TOKEN_PATH), SCOPES)
    if creds.expired and creds.refresh_token:
        creds.refresh(GoogleRequest())
        TOKEN_PATH.write_text(creds.to_json(), encoding="utf-8")
        os.chmod(TOKEN_PATH, 0o600)
    return creds


@app.get("/api/youtube/auth-url")
def youtube_auth_url() -> JSONResponse:
    if not CLIENT_SECRET_PATH.exists():
        return JSONResponse(
            {"error": "Missing client_secret.json in data/credentials."}, status_code=400
        )

    flow = Flow.from_client_secrets_file(
        str(CLIENT_SECRET_PATH),
        scopes=SCOPES,
        redirect_uri="http://localhost:8000/oauth2callback",
    )
    auth_url, state = flow.authorization_url(
        access_type="offline",
        prompt="consent",
        include_granted_scopes="true",
    )
    (CREDENTIALS_DIR / "oauth_state.json").write_text(
        json.dumps({"state": state}), encoding="utf-8"
    )
    return JSONResponse({"auth_url": auth_url})


@app.get("/oauth2callback")
def oauth_callback(request: Request) -> HTMLResponse:
    state_path = CREDENTIALS_DIR / "oauth_state.json"
    if not state_path.exists():
        return HTMLResponse("Missing OAuth state.", status_code=400)

    state = json.loads(state_path.read_text(encoding="utf-8"))["state"]
    flow = Flow.from_client_secrets_file(
        str(CLIENT_SECRET_PATH),
        scopes=SCOPES,
        state=state,
        redirect_uri="http://localhost:8000/oauth2callback",
    )
    flow.fetch_token(authorization_response=str(request.url))
    credentials = flow.credentials
    TOKEN_PATH.write_text(credentials.to_json(), encoding="utf-8")
    os.chmod(TOKEN_PATH, 0o600)
    state_path.unlink(missing_ok=True)
    return HTMLResponse("YouTube authorization complete. You can close this tab.")


@app.post("/api/youtube/upload")
def youtube_upload(
    title: str = Form(...),
    description: str = Form(""),
    tags: str = Form(""),
    visibility: str = Form("private"),
    publish_at: str = Form(""),
    upload_as_short: str = Form(""),
    branding_pack: str = Form(""),
    video_file: UploadFile = File(...),
    thumbnail: UploadFile | None = File(None),
    logo_file: UploadFile | None = File(None),
    logo_position: str = Form(""),
    logo_scale_percent: str = Form(""),
    logo_animated: str = Form(""),
    intro_text: str = Form(""),
    intro_duration_sec: str = Form(""),
    outro_text: str = Form(""),
    outro_duration_sec: str = Form(""),
    subscribe_cta_text: str = Form(""),
    subscribe_cta_duration_sec: str = Form(""),
    subscribe_cta_from_end_sec: str = Form(""),
    end_screen_blocks: str = Form(""),
    end_screen_duration_sec: str = Form(""),
    end_credits_text: str = Form(""),
    end_credits_duration_sec: str = Form(""),
) -> JSONResponse:
    creds = _load_credentials()
    if not creds:
        return JSONResponse({"error": "YouTube not authorized."}, status_code=401)

    service = build("youtube", "v3", credentials=creds)

    video_id = uuid.uuid4().hex
    video_path = OUTPUT_DIR / f"{video_id}_{_safe_filename(video_file.filename)}"
    with video_path.open("wb") as video_buffer:
        shutil.copyfileobj(video_file.file, video_buffer)
    upload_video_path = video_path
    branding_pack_obj = _branding_pack_by_reference(branding_pack)
    if branding_pack_obj is None:
        branding_pack_obj = _default_branding_pack()
    branding_pack_used = ""
    if branding_pack_obj:
        branding_pack_used = str(
            branding_pack_obj.get("id")
            or branding_pack_obj.get("code")
            or branding_pack_obj.get("name")
            or ""
        ).strip()
    logo_applied = False
    logo_error = None
    branding_applied = False
    branding_error = None
    end_credits_applied = False
    end_credits_error = None
    short_mode_enabled = _to_bool(upload_as_short, default=False)
    short_processed = False
    short_error = None
    logo_path: Path | None = None
    request_branding = {
        "logo_position": logo_position,
        "logo_scale_percent": logo_scale_percent,
        "logo_animated": logo_animated,
        "intro_text": intro_text,
        "intro_duration_sec": intro_duration_sec,
        "outro_text": outro_text,
        "outro_duration_sec": outro_duration_sec,
        "subscribe_cta_text": subscribe_cta_text,
        "subscribe_cta_duration_sec": subscribe_cta_duration_sec,
        "subscribe_cta_from_end_sec": subscribe_cta_from_end_sec,
        "end_screen_blocks": end_screen_blocks,
        "end_screen_duration_sec": end_screen_duration_sec,
        "end_credits_text": end_credits_text,
        "end_credits_duration_sec": end_credits_duration_sec,
    }
    if logo_file is not None:
        safe_logo_name = _safe_filename(logo_file.filename or "logo.png")
        logo_path = OUTPUT_DIR / f"{video_id}_logo_{safe_logo_name}"
        with logo_path.open("wb") as logo_buffer:
            shutil.copyfileobj(logo_file.file, logo_buffer)
    elif branding_pack_obj is not None:
        logo_reference = _branding_value(
            request_branding, branding_pack_obj, ("logo_file", "logo_path"), ""
        )
        logo_path = _resolve_batch_asset_path(
            {"logo_file": logo_reference},
            ("logo_file",),
            LOGOS_DIR,
        )

    if logo_path is not None:
        resolved_logo_position = str(
            _branding_value(
                request_branding, branding_pack_obj, ("logo_position",), "top-right"
            )
        ).strip().lower()
        if resolved_logo_position not in {"top-left", "top-right", "bottom-left", "bottom-right", "center"}:
            resolved_logo_position = "top-right"
        resolved_logo_scale = _coerce_int(
            _branding_value(
                request_branding, branding_pack_obj, ("logo_scale_percent",), 15
            ),
            default=15,
            min_value=5,
            max_value=40,
        )
        resolved_logo_animated = _to_bool(
            _branding_value(request_branding, branding_pack_obj, ("logo_animated",), False),
            default=False,
        )
        logo_output_path = OUTPUT_DIR / f"{video_id}_with_logo.mp4"
        try:
            _apply_logo_overlay_to_video(
                video_path=upload_video_path,
                logo_path=logo_path,
                output_path=logo_output_path,
                logo_position=resolved_logo_position,
                logo_scale_percent=resolved_logo_scale,
                logo_animated=resolved_logo_animated,
            )
            upload_video_path = logo_output_path
            logo_applied = True
        except Exception as exc:
            return JSONResponse(
                {
                    "error": "Failed to apply logo overlay.",
                    "details": str(exc),
                },
                status_code=400,
            )

    resolved_intro_text = str(
        _branding_value(request_branding, branding_pack_obj, ("intro_text", "intro_title"), "")
    ).strip()
    resolved_intro_duration = _coerce_int(
        _branding_value(request_branding, branding_pack_obj, ("intro_duration_sec",), 0),
        default=0,
        min_value=0,
        max_value=15,
    )
    resolved_outro_text = str(
        _branding_value(request_branding, branding_pack_obj, ("outro_text", "outro_title"), "")
    ).strip()
    resolved_outro_duration = _coerce_int(
        _branding_value(request_branding, branding_pack_obj, ("outro_duration_sec",), 0),
        default=0,
        min_value=0,
        max_value=20,
    )
    resolved_cta_text = str(
        _branding_value(
            request_branding,
            branding_pack_obj,
            ("subscribe_cta_text", "cta_text", "subscribe_text"),
            "",
        )
    ).strip()
    resolved_cta_duration = _coerce_int(
        _branding_value(
            request_branding,
            branding_pack_obj,
            ("subscribe_cta_duration_sec", "cta_duration_sec"),
            5,
        ),
        default=5,
        min_value=2,
        max_value=20,
    )
    resolved_cta_from_end = _coerce_int(
        _branding_value(
            request_branding,
            branding_pack_obj,
            ("subscribe_cta_from_end_sec", "cta_from_end_sec"),
            12,
        ),
        default=12,
        min_value=2,
        max_value=120,
    )
    resolved_end_screen_blocks = _to_bool(
        _branding_value(request_branding, branding_pack_obj, ("end_screen_blocks",), False),
        default=False,
    )
    resolved_end_screen_duration = _coerce_int(
        _branding_value(request_branding, branding_pack_obj, ("end_screen_duration_sec",), 8),
        default=8,
        min_value=2,
        max_value=20,
    )

    if (
        resolved_intro_text
        or resolved_outro_text
        or resolved_cta_text
        or resolved_end_screen_blocks
    ):
        branding_output_path = OUTPUT_DIR / f"{video_id}_with_branding.mp4"
        try:
            _apply_branding_overlays_to_video(
                video_path=upload_video_path,
                output_path=branding_output_path,
                intro_text=resolved_intro_text,
                intro_duration_sec=resolved_intro_duration,
                outro_text=resolved_outro_text,
                outro_duration_sec=resolved_outro_duration,
                subscribe_cta_text=resolved_cta_text,
                subscribe_cta_duration_sec=resolved_cta_duration,
                subscribe_cta_from_end_sec=resolved_cta_from_end,
                end_screen_blocks=resolved_end_screen_blocks,
                end_screen_duration_sec=resolved_end_screen_duration,
            )
            upload_video_path = branding_output_path
            branding_applied = True
        except Exception as exc:
            branding_error = str(exc)

    resolved_end_credits_text = str(
        _branding_value(
            request_branding,
            branding_pack_obj,
            ("end_credits_text", "credits_text", "outro_text"),
            "",
        )
    ).strip()
    resolved_end_credits_duration = _coerce_int(
        _branding_value(request_branding, branding_pack_obj, ("end_credits_duration_sec",), 5),
        default=5,
        min_value=2,
        max_value=30,
    )
    if resolved_end_credits_text:
        credits_output_path = OUTPUT_DIR / f"{video_id}_with_credits.mp4"
        try:
            _apply_end_credits_to_video(
                video_path=upload_video_path,
                output_path=credits_output_path,
                credits_text=resolved_end_credits_text,
                credits_duration_sec=resolved_end_credits_duration,
            )
            upload_video_path = credits_output_path
            end_credits_applied = True
        except Exception as exc:
            end_credits_error = str(exc)

    if short_mode_enabled:
        shorts_output_path = OUTPUT_DIR / f"{video_id}_shorts.mp4"
        try:
            _prepare_video_for_shorts(
                video_path=upload_video_path,
                output_path=shorts_output_path,
                max_duration_sec=59,
            )
            upload_video_path = shorts_output_path
            short_processed = True
        except Exception as exc:
            short_error = str(exc)
            return JSONResponse(
                {"error": "Failed to prepare Shorts video.", "details": short_error},
                status_code=400,
            )

    effective_title = title
    effective_description = description
    parsed_tags = [tag.strip() for tag in tags.split(",") if tag.strip()]
    if short_mode_enabled:
        if "#shorts" not in effective_title.lower():
            effective_title = f"{effective_title} #Shorts".strip()
        if "#shorts" not in effective_description.lower():
            effective_description = (
                f"{effective_description}\n\n#Shorts".strip()
                if effective_description.strip()
                else "#Shorts"
            )
        if not any(t.lower() in {"shorts", "#shorts"} for t in parsed_tags):
            parsed_tags.append("Shorts")

    safe_visibility = (
        visibility if visibility in ("public", "private", "unlisted") else "private"
    )
    rfc3339_publish_at = _parse_publish_at(publish_at)
    if rfc3339_publish_at and safe_visibility != "private":
        status_dict: dict[str, str] = {
            "privacyStatus": "private",
            "publishAt": rfc3339_publish_at,
        }
    else:
        status_dict = {"privacyStatus": safe_visibility}

    request_body = {
        "snippet": {
            "title": effective_title,
            "description": effective_description,
            "tags": parsed_tags,
            "categoryId": "22",
        },
        "status": status_dict,
    }

    media = MediaFileUpload(str(upload_video_path), resumable=True)
    upload_request = service.videos().insert(
        part=",".join(request_body.keys()),
        body=request_body,
        media_body=media,
    )
    try:
        response = upload_request.execute()
    except HttpError as exc:
        return JSONResponse(
            {
                "error": "YouTube rejected the upload request.",
                "youtube_reason": str(exc),
            },
            status_code=400,
        )
    except Exception as exc:
        return JSONResponse(
            {"error": "Failed to upload video to YouTube.", "details": str(exc)},
            status_code=500,
        )

    thumbnail_source = "none"
    thumbnail_applied = False
    thumbnail_error = None

    if thumbnail is not None:
        thumbnail_path = OUTPUT_DIR / f"{video_id}_thumb_{_safe_filename(thumbnail.filename)}"
        with thumbnail_path.open("wb") as thumb_buffer:
            shutil.copyfileobj(thumbnail.file, thumb_buffer)
        try:
            service.thumbnails().set(
                videoId=response["id"],
                media_body=MediaFileUpload(str(thumbnail_path)),
            ).execute()
            thumbnail_source = "manual"
            thumbnail_applied = True
        except Exception as exc:
            thumbnail_error = str(exc)
    else:
        auto_thumbnail_path = OUTPUT_DIR / f"{video_id}_thumb_auto.jpg"
        try:
            _generate_thumbnail_from_video(video_path, auto_thumbnail_path)
            service.thumbnails().set(
                videoId=response["id"],
                media_body=MediaFileUpload(str(auto_thumbnail_path)),
            ).execute()
            thumbnail_source = "auto"
            thumbnail_applied = True
        except Exception as exc:
            thumbnail_error = str(exc)

    payload = {
        "video_id": response.get("id"),
        "video_url": f"https://www.youtube.com/watch?v={response.get('id')}",
        "thumbnail_applied": thumbnail_applied,
        "thumbnail_source": thumbnail_source,
        "thumbnail_error": thumbnail_error,
        "logo_applied": logo_applied,
        "logo_error": logo_error,
        "branding_applied": branding_applied,
        "branding_error": branding_error,
        "branding_pack": branding_pack_used,
        "end_credits_applied": end_credits_applied,
        "end_credits_error": end_credits_error,
        "short_mode_enabled": short_mode_enabled,
        "short_processed": short_processed,
        "short_error": short_error,
    }
    if rfc3339_publish_at and safe_visibility != "private":
        payload["scheduled_publish_at"] = rfc3339_publish_at
    return JSONResponse(payload)


@app.get("/api/status")
def status() -> JSONResponse:
    return JSONResponse({"time": datetime.utcnow().isoformat()})
